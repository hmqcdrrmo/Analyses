"""
AugHab Flooding Deep-Dive Analysis
==================================

Purpose
-------
Analyze flooding incidents for:
    * 08-10 August 2026
    * 17-19 August 2026

This is deliberately flood-focused. It distinguishes:
    * flood REPORTS (every incident/report row), from
    * UNIQUE FLOODED AREAS (cleaned locations).

Repeated reports from the same cleaned location therefore do not inflate the
number of areas affected.

Main stakeholder outputs
------------------------
For each event:
    - Overview / key metrics
    - Daily summary
    - Barangay distribution
    - District distribution
    - Cluster distribution
    - Depth / severity distribution
    - Barangay x depth matrix
    - District x depth matrix
    - Duration / persistence statistics
    - Start-hour profile
    - 15-minute concurrent flooding timeline
    - Recurring hotspot table
    - Long-duration outlier table
    - Cleaned source reports
    - Stakeholder summary text
    - Several presentation-ready PNG plots

A cross-event comparison workbook is also produced.

Dependencies
------------
    py -m pip install pandas numpy matplotlib openpyxl

Run
---
    py AugHab_flooding_deep_dive.py
"""

from __future__ import annotations

import argparse
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any, Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# USER CONFIGURATION
# =============================================================================

FLOOD_REPORTS_FILE = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis"
    r"\Processed Data\August 2026 Flooding Incidents.xlsx"
)

FLOOD_SHEET = "AUGUST 2026"

OUTPUT_DIRECTORY = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis"
    r"\Outputs\Flooding Deep Dive"
)


@dataclass(frozen=True)
class EventConfig:
    name: str
    label: str
    start_date: pd.Timestamp
    end_date: pd.Timestamp

    @property
    def start_timestamp(self) -> pd.Timestamp:
        return self.start_date.normalize()

    @property
    def end_timestamp(self) -> pd.Timestamp:
        return self.end_date.normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)


EVENTS = (
    EventConfig(
        name="Aug08-10",
        label="08-10 Aug 2026",
        start_date=pd.Timestamp("2026-08-08"),
        end_date=pd.Timestamp("2026-08-10"),
    ),
    EventConfig(
        name="Aug17-19",
        label="17-19 Aug 2026",
        start_date=pd.Timestamp("2026-08-17"),
        end_date=pd.Timestamp("2026-08-19"),
    ),
)

# A 15-minute grid gives a much better picture of concurrent flooding than an
# hourly grid while remaining easy to interpret.
CONCURRENCY_FREQUENCY = "15min"

# Top-N rows retained in plots. Full tables are always written to Excel.
TOP_N_BARANGAYS = 15
TOP_N_HOTSPOTS = 15

# Standard Tukey upper fence for unusually long flood durations.
DURATION_OUTLIER_IQR_MULTIPLIER = 1.5


# =============================================================================
# GENERAL HELPERS
# =============================================================================


def normalize_header(value: Any) -> str:
    text = str(value).strip().casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def find_column(columns: Iterable[Any], *candidates: str) -> Any | None:
    normalized = {normalize_header(c): c for c in columns}

    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized:
            return normalized[key]

    # Conservative contains-match fallback.
    for candidate in candidates:
        key = normalize_header(candidate)
        if not key:
            continue
        for norm, original in normalized.items():
            if key in norm or norm in key:
                return original
    return None


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.casefold().replace("&", " and ")
    text = re.sub(r"\bbarangay\b", "brgy", text)
    text = re.sub(r"\bavenue\b", "ave", text)
    text = re.sub(r"\bstreet\b", "st", text)
    text = re.sub(r"\broad\b", "rd", text)
    text = re.sub(r"\bboulevard\b", "blvd", text)
    text = re.sub(r"\bhighway\b", "hwy", text)
    text = re.sub(r"\s+", " ", text).strip(" ,.;:-")
    return text


def pretty_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def canonicalize_intersection(value: Any) -> str:
    """
    Standardize intersections so:
        'X cor Y' == 'Y cor X'

    Only explicit corner/intersection separators are reordered. Ordinary street
    names containing 'and' are not touched.
    """
    text = normalize_text(value)
    if not text:
        return ""

    text = re.sub(
        r"\s+(?:cor\.?|corner(?:\s+of)?|cnr\.?)\s+",
        " cor ",
        text,
        flags=re.I,
    )
    parts = [p.strip(" ,.;:-") for p in text.split(" cor ") if p.strip()]
    if len(parts) == 2:
        return " cor ".join(sorted(parts))
    return text


def combine_date_time(date_value: Any, time_value: Any) -> pd.Timestamp:
    if pd.isna(date_value):
        return pd.NaT

    date_parsed = pd.to_datetime(date_value, errors="coerce")
    if pd.isna(date_parsed):
        return pd.NaT
    date_parsed = pd.Timestamp(date_parsed).normalize()

    if pd.isna(time_value):
        return date_parsed

    if isinstance(time_value, pd.Timestamp):
        return date_parsed + pd.Timedelta(
            hours=time_value.hour,
            minutes=time_value.minute,
            seconds=time_value.second,
        )

    if isinstance(time_value, datetime):
        return date_parsed + pd.Timedelta(
            hours=time_value.hour,
            minutes=time_value.minute,
            seconds=time_value.second,
        )

    if isinstance(time_value, time):
        return date_parsed + pd.Timedelta(
            hours=time_value.hour,
            minutes=time_value.minute,
            seconds=time_value.second,
        )

    if isinstance(time_value, (float, int, np.floating, np.integer)):
        value = float(time_value)
        # Excel time fraction.
        if 0 <= value < 1:
            return date_parsed + pd.to_timedelta(value, unit="D")

    text = re.sub(r"\s+", " ", str(time_value).strip())
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT

    return date_parsed + pd.Timedelta(
        hours=parsed.hour,
        minutes=parsed.minute,
        seconds=parsed.second,
    )


def parse_duration_minutes(value: Any) -> float:
    if pd.isna(value):
        return np.nan

    if isinstance(value, pd.Timedelta):
        return float(value.total_seconds() / 60)

    if isinstance(value, (int, float, np.integer, np.floating)):
        number = float(value)
        if number < 0:
            return np.nan
        # A small Excel-style fraction is likely a fraction of a day.
        if 0 < number < 1:
            return number * 24 * 60
        return number

    text = str(value).strip().casefold()
    if not text:
        return np.nan

    # HH:MM or HH:MM:SS
    m = re.fullmatch(r"(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?", text)
    if m:
        hours = int(m.group(1))
        minutes = int(m.group(2))
        seconds = int(m.group(3) or 0)
        return hours * 60 + minutes + seconds / 60

    days = 0.0
    hours = 0.0
    minutes = 0.0

    day_match = re.search(r"([\d.]+)\s*(?:day|days|d)\b", text)
    hr_match = re.search(r"([\d.]+)\s*(?:hr|hrs|hour|hours|h)\b", text)
    min_match = re.search(r"([\d.]+)\s*(?:min|mins|minute|minutes|m)\b", text)

    if day_match:
        days = float(day_match.group(1))
    if hr_match:
        hours = float(hr_match.group(1))
    if min_match:
        minutes = float(min_match.group(1))

    if day_match or hr_match or min_match:
        return days * 1440 + hours * 60 + minutes

    try:
        return float(text)
    except ValueError:
        return np.nan


def parse_depth_inches(value: Any) -> float:
    """
    Parse flood depth to inches where possible.

    The source normally contains labels such as:
        Gutter-deep (8 in)
        Half tire-deep (13 in)
        Knee-deep (19 in)
        Tire-deep (26 in)
        Waist-deep (37 in)
        Chest-deep (45 in)
        Head-deep (60 in)

    Explicit parenthetical/numeric inches take precedence.
    """
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)

    text = str(value).strip().casefold()
    if not text:
        return np.nan

    inch_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:in|inch|inches|[\"”])", text)
    if inch_match:
        return float(inch_match.group(1))

    # Fallbacks for common QCDRRMD flood-depth wording.
    mappings = (
        ("above head", 72.0),
        ("head", 60.0),
        ("chest", 45.0),
        ("waist", 37.0),
        ("tire", 26.0),
        ("knee", 19.0),
        ("half tire", 13.0),
        ("half-knee", 10.0),
        ("half knee", 10.0),
        ("gutter", 8.0),
    )
    # Check more specific strings before generic "tire" / "knee".
    for label, inches in sorted(mappings, key=lambda x: len(x[0]), reverse=True):
        if label in text:
            return inches
    return np.nan


def canonical_depth_label(raw_value: Any, inches: float) -> str:
    text = pretty_text(raw_value)
    if text:
        return text
    if pd.isna(inches):
        return "Unknown"
    return f"{inches:g} in"


def duration_band(minutes: float) -> str:
    if pd.isna(minutes):
        return "Unknown"
    if minutes <= 30:
        return "≤30 min"
    if minutes <= 60:
        return "31-60 min"
    if minutes <= 120:
        return "61-120 min"
    if minutes <= 180:
        return "121-180 min"
    return ">180 min"


def severity_band(inches: float) -> str:
    if pd.isna(inches):
        return "Unknown"
    if inches < 19:
        return "Below knee"
    if inches < 26:
        return "Knee-deep"
    if inches < 37:
        return "Tire-deep"
    if inches < 45:
        return "Waist-deep"
    if inches < 60:
        return "Chest-deep"
    return "Head-deep or higher"


def format_hour(timestamp: pd.Timestamp | Any) -> str:
    if pd.isna(timestamp):
        return ""
    return pd.Timestamp(timestamp).strftime("%d %b %Y %H:%M")


def safe_pct(numerator: float, denominator: float) -> float:
    if denominator in (0, None) or pd.isna(denominator):
        return np.nan
    return 100.0 * float(numerator) / float(denominator)


def excel_safe(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime()
    if pd.isna(value) if not isinstance(value, (list, tuple, dict)) else False:
        return None
    return value


# =============================================================================
# INPUT PREPARATION
# =============================================================================


def load_and_clean_flood_reports(path: Path, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"Flood incidents workbook not found:\n{path}")

    frame = pd.read_excel(path, sheet_name=sheet_name, dtype=object, engine="openpyxl")
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all").copy()
    if frame.empty:
        raise ValueError(f"Worksheet '{sheet_name}' contains no usable rows.")

    warnings: list[str] = []

    # Locate fields. Most are optional so the script remains robust to trimmed
    # versions of the workbook; start date/time is required.
    cols = {
        "rain_date": find_column(frame.columns, "DATE (RAIN)", "RAIN DATE"),
        "rain_time": find_column(frame.columns, "TIME (RAIN)", "RAIN TIME"),
        "barangay": find_column(frame.columns, "BARANGAY"),
        "district": find_column(frame.columns, "DISTRICT"),
        "cluster": find_column(frame.columns, "CLUSTER"),
        "street": find_column(frame.columns, "STREET", "LOCATION"),
        "landmark": find_column(
            frame.columns,
            "POINT / LANDMARK / REMARKS",
            "POINT LANDMARK REMARKS",
            "LANDMARK",
            "POINT",
        ),
        "latitude": find_column(frame.columns, "Latitude"),
        "longitude": find_column(frame.columns, "Longitude"),
        "reported_intensity": find_column(frame.columns, "REPORTED RAIN INTENSITY"),
        "depth": find_column(frame.columns, "DEPTH"),
        "start_date": find_column(frame.columns, "FLOOD START DATE"),
        "start_time": find_column(frame.columns, "FLOOD START TIME"),
        "end_date": find_column(frame.columns, "FLOOD END DATE"),
        "end_time": find_column(frame.columns, "FLOOD END TIME"),
        "duration": find_column(frame.columns, "DURATION"),
        "cause": find_column(frame.columns, "CAUSE"),
    }

    if cols["start_date"] is None or cols["start_time"] is None:
        raise ValueError(
            f"Worksheet '{sheet_name}' must contain FLOOD START DATE and FLOOD START TIME."
        )

    if cols["barangay"] is None:
        warnings.append("BARANGAY column was not found; barangay-based tables will be limited.")
    if cols["street"] is None and cols["landmark"] is None:
        warnings.append(
            "No STREET/LOCATION or POINT/LANDMARK field was found. "
            "Unique flooded-area analysis will treat each report as a separate location."
        )
    if cols["district"] is None:
        warnings.append("DISTRICT column was not found; district tables will be blank.")
    if cols["cluster"] is None:
        warnings.append("CLUSTER column was not found; cluster tables will be blank.")
    if cols["depth"] is None:
        warnings.append("DEPTH column was not found; depth/severity statistics will be blank.")

    cleaned = frame.copy()
    cleaned["_Source Row"] = np.arange(2, len(cleaned) + 2)

    def values_or_blank(column: Any | None) -> pd.Series:
        if column is None:
            return pd.Series("", index=cleaned.index, dtype=object)
        return cleaned[column].map(pretty_text)

    cleaned["_Barangay"] = values_or_blank(cols["barangay"]).replace("", "Unknown")
    cleaned["_District"] = values_or_blank(cols["district"]).replace("", "Unknown")
    cleaned["_Cluster"] = values_or_blank(cols["cluster"]).replace("", "Unknown")
    cleaned["_Street"] = values_or_blank(cols["street"])
    cleaned["_Landmark"] = values_or_blank(cols["landmark"])
    cleaned["_Cause"] = values_or_blank(cols["cause"]).replace("", "Unknown")

    if cols["latitude"] is not None:
        cleaned["_Latitude"] = pd.to_numeric(
            cleaned[cols["latitude"]].astype(str).str.replace(",", ".", regex=False),
            errors="coerce",
        )
    else:
        cleaned["_Latitude"] = np.nan

    if cols["longitude"] is not None:
        cleaned["_Longitude"] = pd.to_numeric(
            cleaned[cols["longitude"]].astype(str).str.replace(",", ".", regex=False),
            errors="coerce",
        )
    else:
        cleaned["_Longitude"] = np.nan

    cleaned["_Flood Start"] = [
        combine_date_time(d, t)
        for d, t in zip(cleaned[cols["start_date"]], cleaned[cols["start_time"]])
    ]

    if cols["end_date"] is not None and cols["end_time"] is not None:
        cleaned["_Flood End"] = [
            combine_date_time(d, t)
            for d, t in zip(cleaned[cols["end_date"]], cleaned[cols["end_time"]])
        ]
    else:
        cleaned["_Flood End"] = pd.NaT

    explicit_duration = (
        pd.to_datetime(cleaned["_Flood End"], errors="coerce")
        - pd.to_datetime(cleaned["_Flood Start"], errors="coerce")
    ).dt.total_seconds() / 60.0
    explicit_duration = explicit_duration.where(explicit_duration >= 0)

    if cols["duration"] is not None:
        parsed_duration = cleaned[cols["duration"]].map(parse_duration_minutes)
    else:
        parsed_duration = pd.Series(np.nan, index=cleaned.index, dtype=float)

    cleaned["_Duration Minutes"] = explicit_duration.fillna(parsed_duration)
    cleaned["_Duration Minutes"] = cleaned["_Duration Minutes"].where(
        cleaned["_Duration Minutes"] >= 0
    )

    # Reconstruct end time when a valid duration exists but end datetime is absent.
    reconstruct = (
        cleaned["_Flood End"].isna()
        & cleaned["_Flood Start"].notna()
        & cleaned["_Duration Minutes"].notna()
    )
    cleaned.loc[reconstruct, "_Flood End"] = (
        pd.to_datetime(cleaned.loc[reconstruct, "_Flood Start"])
        + pd.to_timedelta(cleaned.loc[reconstruct, "_Duration Minutes"], unit="m")
    )

    if cols["depth"] is not None:
        cleaned["_Depth Raw"] = cleaned[cols["depth"]].map(pretty_text)
        cleaned["_Depth Inches"] = cleaned[cols["depth"]].map(parse_depth_inches)
    else:
        cleaned["_Depth Raw"] = ""
        cleaned["_Depth Inches"] = np.nan

    cleaned["_Depth Label"] = [
        canonical_depth_label(raw, inches)
        for raw, inches in zip(cleaned["_Depth Raw"], cleaned["_Depth Inches"])
    ]
    cleaned["_Severity Band"] = cleaned["_Depth Inches"].map(severity_band)
    cleaned["_Duration Band"] = cleaned["_Duration Minutes"].map(duration_band)

    # Associated rain timestamp retained for possible future work, but it is not
    # interpreted as a causal lag in this script.
    if cols["rain_date"] is not None and cols["rain_time"] is not None:
        cleaned["_Associated Rain Datetime"] = [
            combine_date_time(d, t)
            for d, t in zip(cleaned[cols["rain_date"]], cleaned[cols["rain_time"]])
        ]
    else:
        cleaned["_Associated Rain Datetime"] = pd.NaT

    # Location cleaning:
    # - standardize intersection order ("X cor Y" == "Y cor X")
    # - include barangay to avoid merging same-named roads across barangays
    # - use landmark as an additional discriminator when available
    location_keys: list[str] = []
    location_displays: list[str] = []

    for idx, row in cleaned.iterrows():
        barangay_key = normalize_text(row["_Barangay"])
        street_key = canonicalize_intersection(row["_Street"])
        landmark_key = normalize_text(row["_Landmark"])

        street_display = pretty_text(row["_Street"])
        landmark_display = pretty_text(row["_Landmark"])

        location_parts = []
        display_parts = []

        if street_key:
            location_parts.append(street_key)
            display_parts.append(street_display)
        if landmark_key and landmark_key not in street_key:
            location_parts.append(landmark_key)
            display_parts.append(landmark_display)

        if not location_parts:
            # Fallback only when the source does not provide usable location text.
            location_parts = [f"report_row_{int(row['_Source Row'])}"]
            display_parts = ["Location not specified"]

        location_key = f"{barangay_key} | " + " | ".join(location_parts)
        location_display = " - ".join([p for p in display_parts if p])

        location_keys.append(location_key)
        location_displays.append(location_display)

    cleaned["_Location Key"] = location_keys
    cleaned["_Location Display"] = location_displays
    cleaned["_Flood Start Hour"] = cleaned["_Flood Start"].dt.floor("h")
    cleaned["_Flood Start Date"] = cleaned["_Flood Start"].dt.floor("D")
    cleaned["_Flood Start Clock Hour"] = cleaned["_Flood Start"].dt.hour

    return cleaned, warnings


# =============================================================================
# EVENT ANALYSIS
# =============================================================================


def select_event_reports(cleaned: pd.DataFrame, event: EventConfig) -> pd.DataFrame:
    valid = cleaned.dropna(subset=["_Flood Start"]).copy()
    mask = (valid["_Flood Start"] >= event.start_timestamp) & (
        valid["_Flood Start"] <= event.end_timestamp
    )
    reports = valid.loc[mask].copy().sort_values("_Flood Start").reset_index(drop=True)

    durations = reports["_Duration Minutes"].dropna()
    if len(durations) >= 4:
        q1 = float(durations.quantile(0.25))
        q3 = float(durations.quantile(0.75))
        iqr = q3 - q1
        upper = q3 + DURATION_OUTLIER_IQR_MULTIPLIER * iqr
    else:
        upper = np.nan

    reports["_Duration Outlier Threshold"] = upper
    reports["_Duration Outlier"] = (
        reports["_Duration Minutes"].notna()
        & np.isfinite(upper)
        & (reports["_Duration Minutes"] > upper)
    )
    return reports


def aggregate_barangays(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    total_reports = len(reports)
    total_unique = reports["_Location Key"].nunique()

    for barangay, group in reports.groupby("_Barangay", dropna=False, sort=False):
        durations = group["_Duration Minutes"].dropna()
        depths = group["_Depth Inches"].dropna()
        unique_areas = group["_Location Key"].nunique()
        rows.append(
            {
                "Barangay": barangay,
                "Flood Reports": len(group),
                "Unique Flooded Areas": unique_areas,
                "Share of Event Reports (%)": safe_pct(len(group), total_reports),
                "Share of Event Unique Areas (%)": safe_pct(unique_areas, total_unique),
                "Affected Days": group["_Flood Start Date"].nunique(),
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "90th Percentile Duration (min)": (
                    durations.quantile(0.90) if not durations.empty else np.nan
                ),
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
                "Maximum Depth (in)": depths.max() if not depths.empty else np.nan,
                "Knee-deep or Higher Reports": int((group["_Depth Inches"] >= 19).sum()),
                "Tire-deep or Higher Reports": int((group["_Depth Inches"] >= 26).sum()),
                "Duration Outliers": int(group["_Duration Outlier"].sum()),
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        ["Unique Flooded Areas", "Flood Reports", "Barangay"],
        ascending=[False, False, True],
    ).reset_index(drop=True)


def aggregate_simple_area(
    reports: pd.DataFrame, group_col: str, display_name: str
) -> pd.DataFrame:
    rows = []
    for key, group in reports.groupby(group_col, dropna=False, sort=False):
        durations = group["_Duration Minutes"].dropna()
        depths = group["_Depth Inches"].dropna()
        rows.append(
            {
                display_name: key,
                "Flood Reports": len(group),
                "Unique Flooded Areas": group["_Location Key"].nunique(),
                "Affected Barangays": group["_Barangay"].nunique(),
                "Affected Days": group["_Flood Start Date"].nunique(),
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
                "Maximum Depth (in)": depths.max() if not depths.empty else np.nan,
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        ["Unique Flooded Areas", "Flood Reports"],
        ascending=[False, False],
    ).reset_index(drop=True)


def depth_distribution(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    total = len(reports)
    for label, group in reports.groupby("_Depth Label", dropna=False, sort=False):
        inches = group["_Depth Inches"].dropna()
        durations = group["_Duration Minutes"].dropna()
        rows.append(
            {
                "Depth": label if label else "Unknown",
                "Representative Depth (in)": (
                    float(inches.median()) if not inches.empty else np.nan
                ),
                "Flood Reports": len(group),
                "Unique Flooded Areas": group["_Location Key"].nunique(),
                "Share of Reports (%)": safe_pct(len(group), total),
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        ["Representative Depth (in)", "Flood Reports"],
        ascending=[True, False],
        na_position="last",
    ).reset_index(drop=True)


def severity_distribution(reports: pd.DataFrame) -> pd.DataFrame:
    order = [
        "Below knee",
        "Knee-deep",
        "Tire-deep",
        "Waist-deep",
        "Chest-deep",
        "Head-deep or higher",
        "Unknown",
    ]
    rows = []
    total = len(reports)
    for label in order:
        group = reports.loc[reports["_Severity Band"] == label]
        if group.empty:
            continue
        rows.append(
            {
                "Severity Band": label,
                "Flood Reports": len(group),
                "Unique Flooded Areas": group["_Location Key"].nunique(),
                "Share of Reports (%)": safe_pct(len(group), total),
            }
        )
    return pd.DataFrame(rows)


def duration_distribution(reports: pd.DataFrame) -> pd.DataFrame:
    order = ["≤30 min", "31-60 min", "61-120 min", "121-180 min", ">180 min", "Unknown"]
    rows = []
    total = len(reports)
    for band in order:
        group = reports.loc[reports["_Duration Band"] == band]
        if group.empty:
            continue
        rows.append(
            {
                "Duration Band": band,
                "Flood Reports": len(group),
                "Unique Flooded Areas": group["_Location Key"].nunique(),
                "Share of Reports (%)": safe_pct(len(group), total),
            }
        )
    return pd.DataFrame(rows)


def duration_by_depth(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for label, group in reports.groupby("_Depth Label", dropna=False, sort=False):
        durations = group["_Duration Minutes"].dropna()
        if durations.empty:
            continue
        depths = group["_Depth Inches"].dropna()
        rows.append(
            {
                "Depth": label if label else "Unknown",
                "Representative Depth (in)": (
                    float(depths.median()) if not depths.empty else np.nan
                ),
                "Reports with Valid Duration": len(durations),
                "Average Duration (min)": durations.mean(),
                "Median Duration (min)": durations.median(),
                "90th Percentile Duration (min)": durations.quantile(0.90),
                "Longest Duration (min)": durations.max(),
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        "Representative Depth (in)", na_position="last"
    ).reset_index(drop=True)


def duration_by_barangay(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for barangay, group in reports.groupby("_Barangay", sort=False):
        durations = group["_Duration Minutes"].dropna()
        if durations.empty:
            continue
        rows.append(
            {
                "Barangay": barangay,
                "Reports with Valid Duration": len(durations),
                "Average Duration (min)": durations.mean(),
                "Median Duration (min)": durations.median(),
                "90th Percentile Duration (min)": durations.quantile(0.90),
                "Longest Duration (min)": durations.max(),
                "Duration Outliers": int(group["_Duration Outlier"].sum()),
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        ["Average Duration (min)", "Reports with Valid Duration"],
        ascending=[False, False],
    ).reset_index(drop=True)


def start_hour_profile(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for hour in range(24):
        group = reports.loc[reports["_Flood Start Clock Hour"] == hour]
        durations = group["_Duration Minutes"].dropna()
        rows.append(
            {
                "Start Hour": hour,
                "Start Hour Label": f"{hour:02d}:00",
                "Flood Reports": len(group),
                "Unique Flooded Areas": group["_Location Key"].nunique(),
                "Affected Barangays": group["_Barangay"].nunique() if not group.empty else 0,
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
            }
        )
    return pd.DataFrame(rows)


def build_concurrency_timeline(
    reports: pd.DataFrame, event: EventConfig
) -> pd.DataFrame:
    timeline = pd.date_range(
        event.start_timestamp,
        event.end_date.normalize() + pd.Timedelta(days=1),
        freq=CONCURRENCY_FREQUENCY,
        inclusive="left",
    )

    valid = reports.dropna(subset=["_Flood Start", "_Flood End"]).copy()
    valid = valid.loc[valid["_Flood End"] >= valid["_Flood Start"]]

    rows = []
    for ts in timeline:
        # Start <= time < end. A report that ends exactly at ts is considered
        # already subsided at ts.
        active = valid.loc[
            (valid["_Flood Start"] <= ts) & (valid["_Flood End"] > ts)
        ]
        rows.append(
            {
                "Date & Time": ts,
                "Active Flood Reports": len(active),
                "Active Unique Flooded Areas": active["_Location Key"].nunique(),
                "Active Barangays": active["_Barangay"].nunique() if not active.empty else 0,
            }
        )
    return pd.DataFrame(rows)


def recurring_hotspots(reports: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for location_key, group in reports.groupby("_Location Key", sort=False):
        durations = group["_Duration Minutes"].dropna()
        depths = group["_Depth Inches"].dropna()

        display = group["_Location Display"].replace("", np.nan).dropna()
        display_name = display.iloc[0] if not display.empty else "Location not specified"

        rows.append(
            {
                "Barangay": group["_Barangay"].iloc[0],
                "District": group["_District"].iloc[0],
                "Cluster": group["_Cluster"].iloc[0],
                "Location": display_name,
                "Location Key": location_key,
                "Flood Reports": len(group),
                "Affected Days": group["_Flood Start Date"].nunique(),
                "First Flood Start": group["_Flood Start"].min(),
                "Last Flood End": group["_Flood End"].max(),
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
                "Maximum Depth (in)": depths.max() if not depths.empty else np.nan,
                "Duration Outliers": int(group["_Duration Outlier"].sum()),
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out

    return out.sort_values(
        ["Flood Reports", "Affected Days", "Maximum Depth (in)"],
        ascending=[False, False, False],
        na_position="last",
    ).reset_index(drop=True)


def daily_summary(
    reports: pd.DataFrame,
    concurrency: pd.DataFrame,
    event: EventConfig,
) -> pd.DataFrame:
    rows = []
    for date in pd.date_range(event.start_date, event.end_date, freq="D"):
        day = reports.loc[reports["_Flood Start Date"] == date]
        durations = day["_Duration Minutes"].dropna()
        depths = day["_Depth Inches"].dropna()

        by_hour = day.groupby("_Flood Start Hour").size()
        if by_hour.empty:
            peak_start_count = 0
            peak_start_hour = pd.NaT
        else:
            peak_start_count = int(by_hour.max())
            peak_start_hour = pd.Timestamp(by_hour.idxmax())

        conc_day = concurrency.loc[concurrency["Date & Time"].dt.floor("D") == date]
        if conc_day.empty:
            peak_concurrent_reports = 0
            peak_concurrent_areas = 0
            peak_concurrent_report_time = pd.NaT
            peak_concurrent_area_time = pd.NaT
        else:
            peak_concurrent_reports = int(conc_day["Active Flood Reports"].max())
            peak_concurrent_areas = int(conc_day["Active Unique Flooded Areas"].max())
            peak_concurrent_report_time = conc_day.loc[
                conc_day["Active Flood Reports"].idxmax(), "Date & Time"
            ]
            peak_concurrent_area_time = conc_day.loc[
                conc_day["Active Unique Flooded Areas"].idxmax(), "Date & Time"
            ]

        rows.append(
            {
                "Date": date,
                "Flood Reports": len(day),
                "Unique Flooded Areas": day["_Location Key"].nunique(),
                "Affected Barangays": day["_Barangay"].nunique() if not day.empty else 0,
                "Affected Districts": (
                    day.loc[day["_District"] != "Unknown", "_District"].nunique()
                    if not day.empty
                    else 0
                ),
                "Average Duration (min)": durations.mean() if not durations.empty else np.nan,
                "Median Duration (min)": durations.median() if not durations.empty else np.nan,
                "90th Percentile Duration (min)": (
                    durations.quantile(0.90) if not durations.empty else np.nan
                ),
                "Longest Duration (min)": durations.max() if not durations.empty else np.nan,
                "Maximum Depth (in)": depths.max() if not depths.empty else np.nan,
                "Knee-deep or Higher Reports": int((day["_Depth Inches"] >= 19).sum()),
                "Tire-deep or Higher Reports": int((day["_Depth Inches"] >= 26).sum()),
                "Peak Flood Starts in One Hour": peak_start_count,
                "Peak Flood Start Hour": peak_start_hour,
                "Peak Concurrent Active Reports": peak_concurrent_reports,
                "Peak Concurrent Reports Time": peak_concurrent_report_time,
                "Peak Concurrent Unique Areas": peak_concurrent_areas,
                "Peak Concurrent Areas Time": peak_concurrent_area_time,
            }
        )
    return pd.DataFrame(rows)


def cross_tab(
    reports: pd.DataFrame, row_col: str, row_name: str
) -> pd.DataFrame:
    if reports.empty:
        return pd.DataFrame()

    data = reports.copy()
    data["_Depth Crosstab"] = data["_Depth Label"].replace("", "Unknown")
    table = pd.crosstab(data[row_col], data["_Depth Crosstab"])
    table["Total"] = table.sum(axis=1)
    table = table.sort_values("Total", ascending=False)
    table.index.name = row_name
    return table.reset_index()


def overview_table(
    reports: pd.DataFrame,
    daily: pd.DataFrame,
    barangays: pd.DataFrame,
    hotspots: pd.DataFrame,
    concurrency: pd.DataFrame,
    event: EventConfig,
) -> pd.DataFrame:
    durations = reports["_Duration Minutes"].dropna()
    depths = reports["_Depth Inches"].dropna()
    unique_areas = reports["_Location Key"].nunique()
    total_reports = len(reports)

    valid_hotspots = hotspots.loc[hotspots["Flood Reports"] > 1] if not hotspots.empty else hotspots
    repeat_reports = (
        int((hotspots["Flood Reports"] - 1).clip(lower=0).sum())
        if not hotspots.empty
        else 0
    )

    if barangays.empty:
        top_barangay_reports = ""
        top_barangay_unique = ""
        top5_share = np.nan
    else:
        top_barangay_reports = barangays.sort_values(
            "Flood Reports", ascending=False
        ).iloc[0]["Barangay"]
        top_barangay_unique = barangays.sort_values(
            "Unique Flooded Areas", ascending=False
        ).iloc[0]["Barangay"]
        top5_share = safe_pct(
            barangays.nlargest(5, "Unique Flooded Areas")["Unique Flooded Areas"].sum(),
            unique_areas,
        )

    if concurrency.empty:
        peak_concurrent_reports = 0
        peak_concurrent_report_time = pd.NaT
        peak_concurrent_areas = 0
        peak_concurrent_area_time = pd.NaT
    else:
        peak_concurrent_reports = int(concurrency["Active Flood Reports"].max())
        peak_concurrent_report_time = concurrency.loc[
            concurrency["Active Flood Reports"].idxmax(), "Date & Time"
        ]
        peak_concurrent_areas = int(concurrency["Active Unique Flooded Areas"].max())
        peak_concurrent_area_time = concurrency.loc[
            concurrency["Active Unique Flooded Areas"].idxmax(), "Date & Time"
        ]

    valid_duration_count = len(durations)
    total_report_hours = durations.sum() / 60 if valid_duration_count else np.nan

    metrics = [
        ("Event", event.label, ""),
        ("Total flood reports", total_reports, "All report rows beginning within the event window"),
        ("Unique flooded areas", unique_areas, "Cleaned unique locations; X cor Y and Y cor X are treated as the same intersection"),
        ("Affected barangays", reports["_Barangay"].nunique() if not reports.empty else 0, ""),
        ("Affected districts", reports.loc[reports["_District"] != "Unknown", "_District"].nunique(), ""),
        ("Affected clusters", reports.loc[reports["_Cluster"] != "Unknown", "_Cluster"].nunique(), ""),
        ("Reports per unique flooded area", total_reports / unique_areas if unique_areas else np.nan, "Values >1 indicate repeat reporting/recurrence"),
        ("Recurring hotspot locations", len(valid_hotspots), "Locations with more than one report during the event"),
        ("Repeat reports beyond first report per location", repeat_reports, ""),
        ("Top 5 barangays' share of unique flooded areas (%)", top5_share, "Concentration indicator"),
        ("Barangay with most flood reports", top_barangay_reports, ""),
        ("Barangay with most unique flooded areas", top_barangay_unique, ""),
        ("Reports with valid duration", valid_duration_count, ""),
        ("Average flood duration (min)", durations.mean() if valid_duration_count else np.nan, ""),
        ("Median flood duration (min)", durations.median() if valid_duration_count else np.nan, ""),
        ("90th percentile flood duration (min)", durations.quantile(0.90) if valid_duration_count else np.nan, ""),
        ("Longest flood duration (min)", durations.max() if valid_duration_count else np.nan, ""),
        ("Reports subsiding within 30 min (%)", safe_pct((durations <= 30).sum(), valid_duration_count), ""),
        ("Reports subsiding within 60 min (%)", safe_pct((durations <= 60).sum(), valid_duration_count), ""),
        ("Reports subsiding within 120 min (%)", safe_pct((durations <= 120).sum(), valid_duration_count), ""),
        ("Reports lasting >180 min (%)", safe_pct((durations > 180).sum(), valid_duration_count), ""),
        ("Reported flood-hours", total_report_hours, "Sum of report durations; this can overlap in time and is not a physical inundated-area measure"),
        ("Maximum reported depth (in)", depths.max() if not depths.empty else np.nan, ""),
        ("Knee-deep or higher reports (%)", safe_pct((reports["_Depth Inches"] >= 19).sum(), total_reports), ""),
        ("Tire-deep or higher reports (%)", safe_pct((reports["_Depth Inches"] >= 26).sum(), total_reports), ""),
        ("Waist-deep or higher reports (%)", safe_pct((reports["_Depth Inches"] >= 37).sum(), total_reports), ""),
        ("Peak concurrent active flood reports", peak_concurrent_reports, ""),
        ("Peak concurrent report time", peak_concurrent_report_time, ""),
        ("Peak concurrent unique flooded areas", peak_concurrent_areas, ""),
        ("Peak concurrent unique-area time", peak_concurrent_area_time, ""),
    ]

    return pd.DataFrame(metrics, columns=["Metric", "Value", "Interpretation / Note"])


# =============================================================================
# PLOTS
# =============================================================================


def save_top_barangays_plot(
    barangays: pd.DataFrame, event: EventConfig, output_dir: Path
) -> None:
    if barangays.empty:
        return

    plot = barangays.head(TOP_N_BARANGAYS).sort_values(
        "Unique Flooded Areas", ascending=True
    )

    fig, ax = plt.subplots(figsize=(11, max(5.5, 0.42 * len(plot) + 1.8)))
    ax.barh(plot["Barangay"], plot["Unique Flooded Areas"])
    ax.set_xlabel("Unique flooded areas")
    ax.set_ylabel("Barangay")
    ax.set_title(f"{event.label}: Barangays by Unique Flooded Areas")
    ax.grid(axis="x", linestyle=":", alpha=0.35)
    fig.tight_layout()
    fig.savefig(output_dir / "barangay_unique_flooded_areas.png", dpi=180)
    plt.close(fig)


def save_daily_plot(daily: pd.DataFrame, event: EventConfig, output_dir: Path) -> None:
    if daily.empty:
        return

    x = np.arange(len(daily))
    width = 0.38

    fig, ax = plt.subplots(figsize=(9.5, 5.5))
    ax.bar(x - width / 2, daily["Flood Reports"], width, label="Flood reports")
    ax.bar(
        x + width / 2,
        daily["Unique Flooded Areas"],
        width,
        label="Unique flooded areas",
    )
    ax.set_xticks(x)
    ax.set_xticklabels(pd.to_datetime(daily["Date"]).dt.strftime("%d %b"))
    ax.set_ylabel("Count")
    ax.set_title(f"{event.label}: Daily Flooding")
    ax.legend()
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    fig.tight_layout()
    fig.savefig(output_dir / "daily_flooding_reports_vs_unique_areas.png", dpi=180)
    plt.close(fig)


def save_depth_plot(
    severity: pd.DataFrame, event: EventConfig, output_dir: Path
) -> None:
    if severity.empty:
        return

    plot = severity.iloc[::-1]
    fig, ax = plt.subplots(figsize=(9.5, 5.7))
    ax.barh(plot["Severity Band"], plot["Flood Reports"])
    ax.set_xlabel("Flood reports")
    ax.set_ylabel("Severity")
    ax.set_title(f"{event.label}: Flood Severity Distribution")
    ax.grid(axis="x", linestyle=":", alpha=0.35)
    fig.tight_layout()
    fig.savefig(output_dir / "flood_severity_distribution.png", dpi=180)
    plt.close(fig)


def save_concurrency_plot(
    concurrency: pd.DataFrame, event: EventConfig, output_dir: Path
) -> None:
    if concurrency.empty:
        return

    fig, ax = plt.subplots(figsize=(14, 5.5))
    ax.plot(
        concurrency["Date & Time"],
        concurrency["Active Unique Flooded Areas"],
        linewidth=1.8,
        label="Active unique flooded areas",
    )
    ax.plot(
        concurrency["Date & Time"],
        concurrency["Active Flood Reports"],
        linewidth=1.2,
        alpha=0.70,
        label="Active flood reports",
    )
    ax.set_ylabel("Concurrent active count")
    ax.set_xlabel("Date & time (PHT)")
    ax.set_title(f"{event.label}: Concurrent Flooding (15-minute intervals)")
    ax.legend()
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(output_dir / "concurrent_flooding_timeline.png", dpi=180)
    plt.close(fig)


def save_hotspot_plot(
    hotspots: pd.DataFrame, event: EventConfig, output_dir: Path
) -> None:
    if hotspots.empty:
        return

    recurring = hotspots.loc[hotspots["Flood Reports"] > 1].head(TOP_N_HOTSPOTS)
    if recurring.empty:
        return

    recurring = recurring.sort_values("Flood Reports", ascending=True).copy()
    labels = recurring["Barangay"].astype(str) + " — " + recurring["Location"].astype(str)

    fig, ax = plt.subplots(
        figsize=(12, max(5.5, 0.48 * len(recurring) + 1.5))
    )
    ax.barh(labels, recurring["Flood Reports"])
    ax.set_xlabel("Flood reports")
    ax.set_ylabel("Recurring location")
    ax.set_title(f"{event.label}: Most Frequently Reported Flooding Hotspots")
    ax.grid(axis="x", linestyle=":", alpha=0.35)
    fig.tight_layout()
    fig.savefig(output_dir / "recurring_flooding_hotspots.png", dpi=180)
    plt.close(fig)


# =============================================================================
# EXCEL OUTPUT
# =============================================================================


def format_excel_workbook(path: Path) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sensible widths, capped so wide text columns do not become enormous.
        for col_idx, column_cells in enumerate(ws.iter_cols(), start=1):
            max_len = 0
            for cell in column_cells[: min(ws.max_row, 250)]:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            width = min(max(max_len + 2, 10), 38)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Basic number formats by header wording.
        headers = {cell.column: str(cell.value or "") for cell in ws[1]}
        for col_idx, header in headers.items():
            h = header.casefold()
            if "(%)" in h or "share" in h:
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.number_format = "0.0"
            elif "(min)" in h or "(in)" in h or "flood-hours" in h:
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.number_format = "0.0"
            elif "date" in h or "time" in h or "start" in h or "end" in h:
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        if isinstance(c.value, datetime):
                            c.number_format = "yyyy-mm-dd hh:mm"

    wb.save(path)


def write_event_workbook(
    path: Path,
    overview: pd.DataFrame,
    daily: pd.DataFrame,
    barangays: pd.DataFrame,
    districts: pd.DataFrame,
    clusters: pd.DataFrame,
    depth: pd.DataFrame,
    severity: pd.DataFrame,
    barangay_depth: pd.DataFrame,
    district_depth: pd.DataFrame,
    duration_barangay: pd.DataFrame,
    duration_depth: pd.DataFrame,
    duration_bands: pd.DataFrame,
    start_profile: pd.DataFrame,
    concurrency: pd.DataFrame,
    hotspots: pd.DataFrame,
    outliers: pd.DataFrame,
    reports: pd.DataFrame,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    # Keep source plus the useful analysis columns in Cleaned Reports.
    source_cols = [c for c in reports.columns if not str(c).startswith("_")]
    analysis_cols = [
        "_Flood Start",
        "_Flood End",
        "_Duration Minutes",
        "_Duration Band",
        "_Depth Inches",
        "_Severity Band",
        "_Barangay",
        "_District",
        "_Cluster",
        "_Location Display",
        "_Location Key",
        "_Duration Outlier",
    ]
    clean_export_cols = source_cols + [c for c in analysis_cols if c in reports.columns]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        daily.to_excel(writer, sheet_name="Daily Summary", index=False)
        barangays.to_excel(writer, sheet_name="Barangay Distribution", index=False)
        districts.to_excel(writer, sheet_name="District Distribution", index=False)
        clusters.to_excel(writer, sheet_name="Cluster Distribution", index=False)
        depth.to_excel(writer, sheet_name="Depth Distribution", index=False)
        severity.to_excel(writer, sheet_name="Severity Bands", index=False)
        barangay_depth.to_excel(writer, sheet_name="Barangay x Depth", index=False)
        district_depth.to_excel(writer, sheet_name="District x Depth", index=False)
        duration_barangay.to_excel(writer, sheet_name="Duration by Barangay", index=False)
        duration_depth.to_excel(writer, sheet_name="Duration by Depth", index=False)
        duration_bands.to_excel(writer, sheet_name="Duration Bands", index=False)
        start_profile.to_excel(writer, sheet_name="Start Hour Profile", index=False)
        concurrency.to_excel(writer, sheet_name="Concurrent Flooding", index=False)
        hotspots.to_excel(writer, sheet_name="Recurring Hotspots", index=False)
        outliers.to_excel(writer, sheet_name="Duration Outliers", index=False)
        reports[clean_export_cols].to_excel(writer, sheet_name="Cleaned Reports", index=False)

    format_excel_workbook(path)


# =============================================================================
# STAKEHOLDER SUMMARY
# =============================================================================


def build_stakeholder_summary(
    event: EventConfig,
    overview: pd.DataFrame,
    daily: pd.DataFrame,
    barangays: pd.DataFrame,
    hotspots: pd.DataFrame,
) -> str:
    metric = dict(zip(overview["Metric"], overview["Value"]))

    lines = [
        f"{event.label} FLOODING DEEP-DIVE",
        "=" * (len(event.label) + 20),
        "",
        "EVENT SCALE",
        f"- Flood reports: {int(metric.get('Total flood reports', 0) or 0)}",
        f"- Unique flooded areas: {int(metric.get('Unique flooded areas', 0) or 0)}",
        f"- Affected barangays: {int(metric.get('Affected barangays', 0) or 0)}",
    ]

    if not barangays.empty:
        top_unique = barangays.iloc[0]
        top_reports = barangays.sort_values("Flood Reports", ascending=False).iloc[0]
        lines += [
            "",
            "SPATIAL CONCENTRATION",
            f"- Most unique flooded areas: {top_unique['Barangay']} "
            f"({int(top_unique['Unique Flooded Areas'])} areas; "
            f"{int(top_unique['Flood Reports'])} reports).",
            f"- Most flood reports: {top_reports['Barangay']} "
            f"({int(top_reports['Flood Reports'])} reports).",
        ]
        top5 = metric.get("Top 5 barangays' share of unique flooded areas (%)", np.nan)
        if not pd.isna(top5):
            lines.append(
                f"- The top five barangays accounted for {float(top5):.1f}% "
                "of all unique flooded areas."
            )

    avg = metric.get("Average flood duration (min)", np.nan)
    med = metric.get("Median flood duration (min)", np.nan)
    p90 = metric.get("90th percentile flood duration (min)", np.nan)
    longest = metric.get("Longest flood duration (min)", np.nan)

    lines += ["", "PERSISTENCE"]
    if not pd.isna(avg):
        lines.append(f"- Average flood duration: {float(avg):.0f} minutes.")
    if not pd.isna(med):
        lines.append(f"- Median flood duration: {float(med):.0f} minutes.")
    if not pd.isna(p90):
        lines.append(f"- 90% of valid-duration reports lasted about {float(p90):.0f} minutes or less.")
    if not pd.isna(longest):
        lines.append(f"- Longest recorded flood duration: {float(longest):.0f} minutes.")

    peak_active = metric.get("Peak concurrent unique flooded areas", np.nan)
    peak_active_time = metric.get("Peak concurrent unique-area time", pd.NaT)
    if not pd.isna(peak_active):
        lines += [
            "",
            "SIMULTANEOUS FLOODING",
            f"- Peak concurrent unique flooded areas: {int(peak_active)} "
            f"at {format_hour(peak_active_time)}.",
        ]

    recurring = hotspots.loc[hotspots["Flood Reports"] > 1] if not hotspots.empty else hotspots
    lines += [
        "",
        "RECURRENCE / HOTSPOTS",
        f"- Recurring hotspot locations: {len(recurring)}.",
    ]
    if not recurring.empty:
        top = recurring.iloc[0]
        lines.append(
            f"- Most repeatedly reported location: {top['Barangay']} — {top['Location']} "
            f"({int(top['Flood Reports'])} reports across {int(top['Affected Days'])} day(s))."
        )

    if not daily.empty:
        peak_day = daily.loc[daily["Unique Flooded Areas"].idxmax()]
        lines += [
            "",
            "DAILY PEAK",
            f"- Greatest daily spatial extent by report locations: "
            f"{pd.Timestamp(peak_day['Date']).strftime('%d %b %Y')} "
            f"with {int(peak_day['Unique Flooded Areas'])} unique flooded areas "
            f"and {int(peak_day['Flood Reports'])} reports.",
        ]

    lines += [
        "",
        "INTERPRETATION NOTES",
        "- 'Flood reports' counts every report row.",
        "- 'Unique flooded areas' counts cleaned locations, reducing inflation from repeated updates.",
        "- Intersections written as 'X cor Y' and 'Y cor X' are treated as one location.",
        "- 'Reported flood-hours' is the sum of report durations. Concurrent reports can overlap, "
        "so it is a workload/persistence indicator rather than a physical inundated-area measure.",
        "- Concurrent flooding uses only reports with usable start and end times.",
    ]

    return "\n".join(lines)


# =============================================================================
# EVENT COMPARISON
# =============================================================================


def make_event_comparison(
    event_results: dict[str, dict[str, Any]], output_dir: Path
) -> Path:
    rows = []
    daily_frames = []
    barangay_frames = []
    depth_frames = []

    for event in EVENTS:
        result = event_results[event.name]
        overview = result["overview"]
        metric = dict(zip(overview["Metric"], overview["Value"]))

        rows.append(
            {
                "Event": event.label,
                "Flood Reports": metric.get("Total flood reports"),
                "Unique Flooded Areas": metric.get("Unique flooded areas"),
                "Affected Barangays": metric.get("Affected barangays"),
                "Recurring Hotspot Locations": metric.get("Recurring hotspot locations"),
                "Average Duration (min)": metric.get("Average flood duration (min)"),
                "Median Duration (min)": metric.get("Median flood duration (min)"),
                "90th Percentile Duration (min)": metric.get("90th percentile flood duration (min)"),
                "Longest Duration (min)": metric.get("Longest flood duration (min)"),
                "Reported Flood-Hours": metric.get("Reported flood-hours"),
                "Knee-deep or Higher Reports (%)": metric.get("Knee-deep or higher reports (%)"),
                "Tire-deep or Higher Reports (%)": metric.get("Tire-deep or higher reports (%)"),
                "Peak Concurrent Unique Areas": metric.get("Peak concurrent unique flooded areas"),
                "Peak Concurrent Unique-Area Time": metric.get("Peak concurrent unique-area time"),
                "Top 5 Barangay Concentration (%)": metric.get(
                    "Top 5 barangays' share of unique flooded areas (%)"
                ),
            }
        )

        daily = result["daily"].copy()
        daily.insert(0, "Event", event.label)
        daily_frames.append(daily)

        barangay = result["barangays"][
            ["Barangay", "Flood Reports", "Unique Flooded Areas"]
        ].copy()
        barangay.insert(0, "Event", event.label)
        barangay_frames.append(barangay)

        depth = result["severity"].copy()
        depth.insert(0, "Event", event.label)
        depth_frames.append(depth)

    comparison = pd.DataFrame(rows)
    daily_all = pd.concat(daily_frames, ignore_index=True) if daily_frames else pd.DataFrame()
    barangay_all = (
        pd.concat(barangay_frames, ignore_index=True) if barangay_frames else pd.DataFrame()
    )
    depth_all = pd.concat(depth_frames, ignore_index=True) if depth_frames else pd.DataFrame()

    # Wide barangay comparison.
    if not barangay_all.empty:
        barangay_wide = barangay_all.pivot_table(
            index="Barangay",
            columns="Event",
            values=["Flood Reports", "Unique Flooded Areas"],
            aggfunc="sum",
            fill_value=0,
        )
        barangay_wide.columns = [
            f"{metric} - {event}" for metric, event in barangay_wide.columns
        ]
        barangay_wide = barangay_wide.reset_index()
    else:
        barangay_wide = pd.DataFrame()

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "AugHab_flooding_event_comparison.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        comparison.to_excel(writer, sheet_name="Event Comparison", index=False)
        daily_all.to_excel(writer, sheet_name="Daily Comparison", index=False)
        barangay_wide.to_excel(writer, sheet_name="Barangay Comparison", index=False)
        depth_all.to_excel(writer, sheet_name="Severity Comparison", index=False)

    format_excel_workbook(path)
    return path


# =============================================================================
# MAIN PROCESSING
# =============================================================================


def analyze_event(
    event: EventConfig,
    all_reports: pd.DataFrame,
    output_root: Path,
    warnings: list[str],
) -> dict[str, Any]:
    event_dir = output_root / event.name
    plots_dir = event_dir / "plots"
    event_dir.mkdir(parents=True, exist_ok=True)
    plots_dir.mkdir(parents=True, exist_ok=True)

    reports = select_event_reports(all_reports, event)

    barangays = aggregate_barangays(reports)
    districts = aggregate_simple_area(reports, "_District", "District")
    clusters = aggregate_simple_area(reports, "_Cluster", "Cluster")
    depth = depth_distribution(reports)
    severity = severity_distribution(reports)
    duration_bands = duration_distribution(reports)
    duration_depth = duration_by_depth(reports)
    duration_barangay = duration_by_barangay(reports)
    start_profile = start_hour_profile(reports)
    concurrency = build_concurrency_timeline(reports, event)
    hotspots = recurring_hotspots(reports)
    daily = daily_summary(reports, concurrency, event)

    barangay_depth = cross_tab(reports, "_Barangay", "Barangay")
    district_depth = cross_tab(reports, "_District", "District")

    overview = overview_table(
        reports, daily, barangays, hotspots, concurrency, event
    )

    outlier_export_cols = [
        "_Barangay",
        "_District",
        "_Cluster",
        "_Location Display",
        "_Flood Start",
        "_Flood End",
        "_Duration Minutes",
        "_Depth Raw",
        "_Depth Inches",
        "_Duration Outlier Threshold",
    ]
    outliers = reports.loc[reports["_Duration Outlier"], outlier_export_cols].copy()
    outliers.columns = [
        "Barangay",
        "District",
        "Cluster",
        "Location",
        "Flood Start",
        "Flood End",
        "Duration (min)",
        "Depth",
        "Depth (in)",
        "Outlier Threshold (min)",
    ]

    workbook_path = event_dir / "flooding_deep_dive.xlsx"
    write_event_workbook(
        workbook_path,
        overview,
        daily,
        barangays,
        districts,
        clusters,
        depth,
        severity,
        barangay_depth,
        district_depth,
        duration_barangay,
        duration_depth,
        duration_bands,
        start_profile,
        concurrency,
        hotspots,
        outliers,
        reports,
    )

    summary = build_stakeholder_summary(
        event, overview, daily, barangays, hotspots
    )

    if warnings:
        summary += "\n\nDATA / STRUCTURE WARNINGS\n" + "\n".join(
            f"- {warning}" for warning in warnings
        )

    (event_dir / "stakeholder_summary.txt").write_text(
        summary, encoding="utf-8-sig"
    )

    save_top_barangays_plot(barangays, event, plots_dir)
    save_daily_plot(daily, event, plots_dir)
    save_depth_plot(severity, event, plots_dir)
    save_concurrency_plot(concurrency, event, plots_dir)
    save_hotspot_plot(hotspots, event, plots_dir)

    return {
        "reports": reports,
        "overview": overview,
        "daily": daily,
        "barangays": barangays,
        "districts": districts,
        "clusters": clusters,
        "depth": depth,
        "severity": severity,
        "duration_bands": duration_bands,
        "duration_depth": duration_depth,
        "duration_barangay": duration_barangay,
        "start_profile": start_profile,
        "concurrency": concurrency,
        "hotspots": hotspots,
        "workbook": workbook_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Deep-dive analysis of August 2026 flooding incidents."
    )
    parser.add_argument(
        "--flood-reports",
        type=Path,
        default=FLOOD_REPORTS_FILE,
        help="Path to August 2026 Flooding Incidents.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=FLOOD_SHEET,
        help="Worksheet containing the flood incidents.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_DIRECTORY,
        help="Output directory.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    all_reports, warnings = load_and_clean_flood_reports(
        args.flood_reports, args.sheet
    )

    invalid_start = int(all_reports["_Flood Start"].isna().sum())
    if invalid_start:
        warnings = warnings + [
            f"{invalid_start} source row(s) had an unusable flood start date/time "
            "and could not be assigned to either event."
        ]

    event_results: dict[str, dict[str, Any]] = {}

    # The events are independent; keeping the event loop simple makes the output
    # easy to audit and avoids hidden shared state.
    for event in EVENTS:
        print(f"Processing {event.label}...")
        event_results[event.name] = analyze_event(
            event, all_reports, args.output, warnings
        )

    comparison_path = make_event_comparison(event_results, args.output)

    print("")
    print("Flooding deep-dive complete.")
    for event in EVENTS:
        print(f"  {event.label}: {event_results[event.name]['workbook']}")
    print(f"  Comparison: {comparison_path}")
    print(f"  Output root: {args.output}")


if __name__ == "__main__":
    main()
