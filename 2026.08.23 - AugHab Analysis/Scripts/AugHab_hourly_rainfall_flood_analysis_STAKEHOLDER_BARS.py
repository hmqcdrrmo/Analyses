r"""
Analyze and compare the 08-10 August 2026 and 17-19 August 2026
Quezon City Habagat rainfall/flooding events.

Main outputs, separated by event:
    <Outputs>\Aug08-10\
    <Outputs>\Aug17-19\

Each event folder contains:
    plots\hourly_rainfall_3day.png
    plots\hourly_rainfall_and_flood_reports_3day.png
    plots\hourly_rainfall_and_flood_duration_3day.png  (horizontal duration bars)
    tables\cleaned_rainfall_hourly.xlsx
    tables\flood_reports_hourly.xlsx
    tables\event_summary.xlsx
    qc\rejected_rainfall_values.csv
    qc\sensor_status.csv
    qc\qc_summary.txt
    event_summary.txt

Quality-control approach
------------------------
1. Missing/non-numeric values remain NaN and are excluded from calculations.
2. Negative rainfall and values above the hard hourly ceiling are rejected.
3. Likely dead/stuck sensors are excluded from network statistics.
4. An additional automatic network-outlier rejection is deliberately
   conservative. A value must be very large AND much larger than the rest of
   the network AND extremely unusual by robust statistics before it is removed.
   This is intended to catch values such as 800 mm/h among 50-100 mm/h gauges
   without deleting legitimate localized heavy/torrential rainfall.
5. Daily rainfall is the mean of station daily accumulations based on whatever
   valid hourly observations are available for each station-day. Missing cells
   are skipped, never converted to zero; a station-day contributes as long as it
   has at least one valid hourly observation. Hourly rainfall is the mean of all
   valid, included gauges reporting in that hour.

Dependencies:
    py -m pip install pandas numpy matplotlib openpyxl

Run:
    py AugHab_hourly_rainfall_flood_analysis.py
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    from matplotlib.patches import Patch
    from matplotlib.ticker import MaxNLocator
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency. Run: py -m pip install pandas numpy matplotlib openpyxl"
    ) from exc


# =============================================================================
# USER CONFIGURATION
# =============================================================================

RF_0810_FILE = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis"
    r"\Processed Data\RF_0810.xlsx"
)

RF_1719_FILE = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis"
    r"\Processed Data\RF_1719.xlsx"
)

FLOOD_REPORTS_FILE = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis"
    r"\Processed Data\August 2026 Flooding Incidents.xlsx"
)

FLOOD_SHEET = "AUGUST 2026"

OUTPUT_DIRECTORY = Path(
    r"C:\Users\QCUSER\Documents\Analysis\2026.08.23 - AugHab Analysis\Outputs"
)


@dataclass(frozen=True)
class EventConfig:
    name: str
    label: str
    rainfall_file: Path
    start_date: pd.Timestamp
    end_date: pd.Timestamp
    color: str

    @property
    def start_timestamp(self) -> pd.Timestamp:
        return self.start_date.floor("D")

    @property
    def end_timestamp(self) -> pd.Timestamp:
        return self.end_date.floor("D") + pd.Timedelta(hours=23)

    @property
    def full_hourly_index(self) -> pd.DatetimeIndex:
        return pd.date_range(self.start_timestamp, self.end_timestamp, freq="h")


EVENTS = (
    EventConfig(
        name="Aug08-10",
        label="08-10 Aug 2026",
        rainfall_file=RF_0810_FILE,
        start_date=pd.Timestamp("2026-08-08"),
        end_date=pd.Timestamp("2026-08-10"),
        color="#2878B5",
    ),
    EventConfig(
        name="Aug17-19",
        label="17-19 Aug 2026",
        rainfall_file=RF_1719_FILE,
        start_date=pd.Timestamp("2026-08-17"),
        end_date=pd.Timestamp("2026-08-19"),
        color="#E07A1F",
    ),
)


@dataclass(frozen=True)
class QCSettings:
    # Objectively invalid / impossible for this workflow.
    hard_max_hourly_mm: float = 300.0

    # Conservative automatic network-outlier rejection. ALL core conditions
    # must be satisfied before a value below the hard ceiling is removed.
    obvious_outlier_min_mm: float = 150.0
    obvious_outlier_min_ratio_to_network_median: float = 3.0
    obvious_outlier_min_difference_mm: float = 100.0
    obvious_outlier_modified_z: float = 8.0
    obvious_outlier_iqr_multiplier: float = 6.0
    minimum_reporting_sensors_for_outlier_test: int = 8

    # Sensor-status screening.
    network_wet_threshold_mm: float = 2.5
    minimum_network_wet_hours: int = 3
    inactive_zero_fraction: float = 0.90
    constant_sensor_minimum_coverage: float = 0.90

    # Daily accumulation from available observations. A station-day is used
    # whenever it has at least this many valid hourly values. Missing hours are
    # skipped rather than treated as zero.
    minimum_valid_hours_for_daily_accumulation: int = 1


DEFAULT_QC = QCSettings()

# Requested reference lines. Torrential was not assigned a color in the prompt;
# purple is used here to remain distinct from Intense (brown).
HOURLY_INTENSITY_LINES = (
    (2.5, "Moderate", "#F39C12"),
    (7.5, "Heavy", "#D62728"),
    (15.0, "Intense", "#8B4513"),
    (30.0, "Torrential", "#7B2CBF"),
)

REJECTED_COLUMNS = [
    "event",
    "timestamp",
    "station",
    "original_value",
    "reason",
    "network_median_mm",
    "modified_z",
    "network_ratio",
    "network_difference_mm",
    "iqr_extreme_upper_mm",
]


# =============================================================================
# PARSING / GENERAL HELPERS
# =============================================================================


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").strip())


def normalize_header(value: Any) -> str:
    text = clean_text(value).casefold().replace("_", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def parse_datetime_cell(value: Any) -> pd.Timestamp:
    if pd.isna(value):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value
    if isinstance(value, datetime):
        return pd.Timestamp(value)
    if isinstance(value, date):
        return pd.Timestamp(value)
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            return pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")
        except (ValueError, TypeError, OverflowError):
            return pd.NaT

    text = clean_text(value)
    for fmt in (
        "%d %m %Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%b %d, %Y %I:%M %p",
        "%B %d, %Y %I:%M %p",
    ):
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            pass
    return pd.to_datetime(text, dayfirst=True, errors="coerce")


def parse_rainfall_cell(value: Any) -> float:
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = clean_text(value).replace(" ", "")
    if text.casefold() in {"", "-", "--", "na", "n/a", "nan", "none", "null"}:
        return np.nan
    # Rainfall workbooks may contain decimal commas.
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return np.nan


def meaningful_unparsed_value(value: Any) -> bool:
    if pd.isna(value):
        return False
    return clean_text(value).casefold() not in {
        "",
        "-",
        "--",
        "na",
        "n/a",
        "nan",
        "none",
        "null",
    }


def find_datetime_column(columns: Iterable[Any]) -> Any:
    columns = list(columns)
    normalized = {normalize_header(c): c for c in columns}
    for candidate in (
        "date time",
        "date and time",
        "datetime",
        "date pht",
        "date",
    ):
        if candidate in normalized:
            return normalized[candidate]
    return columns[0]


def find_column(columns: Iterable[Any], *candidates: str) -> Any | None:
    normalized = {normalize_header(c): c for c in columns}
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized:
            return normalized[key]
    return None


def combine_excel_date_time(date_value: Any, time_value: Any) -> pd.Timestamp:
    """Combine Excel/string date and time cells into one timestamp."""
    if pd.isna(date_value):
        return pd.NaT

    date_ts = parse_datetime_cell(date_value)
    if pd.isna(date_ts):
        return pd.NaT
    date_ts = date_ts.floor("D")

    if pd.isna(time_value):
        return date_ts

    if isinstance(time_value, time):
        return date_ts + pd.Timedelta(
            hours=time_value.hour,
            minutes=time_value.minute,
            seconds=time_value.second,
        )
    if isinstance(time_value, (datetime, pd.Timestamp)):
        t = pd.Timestamp(time_value)
        return date_ts + pd.Timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    if isinstance(time_value, (int, float, np.integer, np.floating)):
        number = float(time_value)
        # Excel time is normally a fraction of a day. If a full serial appears,
        # keep only the fractional day.
        fraction = number % 1.0
        return date_ts + pd.to_timedelta(fraction, unit="D")

    text = clean_text(time_value)
    for fmt in ("%I:%M %p", "%I:%M:%S %p", "%H:%M", "%H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            return date_ts + pd.Timedelta(
                hours=parsed.hour, minutes=parsed.minute, seconds=parsed.second
            )
        except ValueError:
            pass

    combined = pd.to_datetime(
        f"{date_ts:%Y-%m-%d} {text}",
        errors="coerce",
    )
    return pd.Timestamp(combined) if not pd.isna(combined) else pd.NaT


def parse_duration_minutes(value: Any) -> float:
    """Parse common Excel/text flood-duration values to minutes."""
    if pd.isna(value):
        return np.nan
    if isinstance(value, pd.Timedelta):
        return float(value.total_seconds() / 60.0)
    if isinstance(value, time):
        return float(value.hour * 60 + value.minute + value.second / 60.0)
    if isinstance(value, (int, float, np.integer, np.floating)):
        number = float(value)
        if number < 0:
            return np.nan
        # Excel duration/time values are commonly stored as fractions of a day.
        if number < 1:
            return number * 24.0 * 60.0
        # Otherwise assume an explicit number of minutes.
        return number

    text = clean_text(value).casefold()
    if text in {"", "-", "--", "na", "n/a", "nan", "none"}:
        return np.nan

    # HH:MM or H:MM duration form.
    match = re.fullmatch(r"(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        return float(hours * 60 + minutes + seconds / 60.0)

    hours = 0.0
    minutes = 0.0
    hour_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:h|hr|hrs|hour|hours)\b", text)
    minute_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:m|min|mins|minute|minutes)\b", text)
    if hour_match:
        hours = float(hour_match.group(1))
    if minute_match:
        minutes = float(minute_match.group(1))
    if hour_match or minute_match:
        return hours * 60.0 + minutes

    # Bare numeric text: assume minutes.
    try:
        number = float(text.replace(",", "."))
        return number if number >= 0 else np.nan
    except ValueError:
        return np.nan


def mean_datetime(values: pd.Series) -> pd.Timestamp:
    """Arithmetic mean of valid datetimes, returned as a timestamp."""
    valid = pd.to_datetime(values, errors="coerce").dropna()
    if valid.empty:
        return pd.NaT
    return pd.Timestamp(valid.astype("int64").mean())


def robust_distribution_metrics(values: Iterable[float]) -> dict[str, float]:
    series = pd.Series(values, dtype=float).dropna()
    if series.empty:
        return {"median": np.nan, "mad": np.nan, "q1": np.nan, "q3": np.nan, "iqr": np.nan}
    median = float(series.median())
    mad = float((series - median).abs().median())
    q1 = float(series.quantile(0.25))
    q3 = float(series.quantile(0.75))
    return {"median": median, "mad": mad, "q1": q1, "q3": q3, "iqr": q3 - q1}


def modified_z(value: float, median: float, mad: float) -> float:
    if pd.isna(value) or pd.isna(median) or pd.isna(mad):
        return np.nan
    if mad == 0:
        if value > median:
            return np.inf
        if value < median:
            return -np.inf
        return 0.0
    return 0.6745 * (value - median) / mad


def common_axis_upper(values: Iterable[float], minimum_peak: float = 0.0, padding: float = 1.15) -> float:
    series = pd.Series(values, dtype=float).replace([np.inf, -np.inf], np.nan).dropna()
    observed = float(series.max()) if not series.empty else 0.0
    target = max(observed, minimum_peak)
    if target <= 0:
        return 1.0
    padded = target * padding
    ticks = MaxNLocator(nbins=6, min_n_ticks=4).tick_values(0, padded)
    eligible = ticks[ticks >= padded]
    return float(eligible[0] if len(eligible) else ticks[-1])


def safe_excel_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def format_excel_workbook(path: Path) -> None:
    """Light formatting for every worksheet after pandas writes the workbook."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column_cells in ws.columns:
            values = [clean_text(cell.value) for cell in column_cells[: min(ws.max_row, 200)]]
            width = min(max([len(v) for v in values] + [10]) + 2, 42)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(path)


def save_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out = frame.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%Y-%m-%d %H:%M")
    out.to_csv(path, index=False, encoding="utf-8-sig", float_format="%.3f")


# =============================================================================
# RAINFALL QC AND EVENT PROCESSING
# =============================================================================


def load_rainfall_event(
    event: EventConfig,
    settings: QCSettings,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    if not event.rainfall_file.exists():
        raise FileNotFoundError(f"Rainfall workbook not found for {event.name}:\n{event.rainfall_file}")

    raw = pd.read_excel(event.rainfall_file, sheet_name=0, dtype=object, engine="openpyxl")
    raw = raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if raw.empty or raw.shape[1] < 2:
        raise ValueError(f"{event.rainfall_file.name} has no usable rainfall table.")

    datetime_col = find_datetime_column(raw.columns)
    parsed_time = raw[datetime_col].map(parse_datetime_cell)
    invalid_timestamp_rows = int(parsed_time.isna().sum())

    valid = parsed_time.notna()
    raw = raw.loc[valid].copy()
    parsed_time = parsed_time.loc[valid]
    in_event = (parsed_time >= event.start_timestamp) & (parsed_time <= event.end_timestamp)
    outside_event_rows = int((~in_event).sum())
    raw = raw.loc[in_event].copy()
    parsed_time = parsed_time.loc[in_event]
    if raw.empty:
        raise ValueError(
            f"No rainfall timestamps in {event.rainfall_file.name} fall within {event.label}."
        )

    station_cols = [c for c in raw.columns if c != datetime_col]
    numeric = pd.DataFrame(index=raw.index)
    rejected: list[dict[str, Any]] = []
    malformed_count = 0

    for original_station in station_cols:
        station = clean_text(original_station)
        parsed = raw[original_station].map(parse_rainfall_cell)
        numeric[station] = parsed
        malformed = parsed.isna() & raw[original_station].map(meaningful_unparsed_value)
        malformed_count += int(malformed.sum())
        for idx in raw.index[malformed]:
            rejected.append(
                {
                    "event": event.name,
                    "timestamp": parsed_time.loc[idx],
                    "station": station,
                    "original_value": raw.at[idx, original_station],
                    "reason": "non_numeric_value",
                    "network_median_mm": np.nan,
                    "modified_z": np.nan,
                    "network_ratio": np.nan,
                    "network_difference_mm": np.nan,
                    "iqr_extreme_upper_mm": np.nan,
                }
            )

    # Hard rejection before any network statistics.
    for idx, timestamp in parsed_time.items():
        for station in numeric.columns:
            value = numeric.at[idx, station]
            if pd.isna(value):
                continue
            reason = None
            if value < 0:
                reason = "negative_rainfall"
            elif value > settings.hard_max_hourly_mm:
                reason = f"above_hard_limit_{settings.hard_max_hourly_mm:g}_mm"
            if reason:
                rejected.append(
                    {
                        "event": event.name,
                        "timestamp": timestamp,
                        "station": station,
                        "original_value": value,
                        "reason": reason,
                        "network_median_mm": np.nan,
                        "modified_z": np.nan,
                        "network_ratio": np.nan,
                        "network_difference_mm": np.nan,
                        "iqr_extreme_upper_mm": np.nan,
                    }
                )
                numeric.at[idx, station] = np.nan

    numeric.index = pd.DatetimeIndex(parsed_time.values, name="Date & Time")
    duplicate_rows = int(numeric.index.duplicated(keep=False).sum())
    if duplicate_rows:
        numeric = numeric.groupby(level=0).mean()

    numeric = numeric.sort_index().reindex(event.full_hourly_index)
    numeric.index.name = "Date & Time"

    rejected_frame = pd.DataFrame(rejected, columns=REJECTED_COLUMNS)
    metadata = {
        "invalid_timestamp_rows_removed": invalid_timestamp_rows,
        "rows_outside_event_removed": outside_event_rows,
        "duplicate_timestamp_rows_collapsed": duplicate_rows,
        "malformed_rainfall_cells_rejected": malformed_count,
        "missing_cells_after_reindex_before_sensor_qc": int(numeric.isna().sum().sum()),
        "input_station_columns": int(numeric.shape[1]),
    }
    return numeric, rejected_frame, metadata


def identify_sensor_status(
    event: EventConfig,
    rainfall: pd.DataFrame,
    settings: QCSettings,
) -> pd.DataFrame:
    """Exclude all-missing, clearly inactive, and clearly stuck sensors."""
    candidate_columns = [c for c in rainfall.columns if rainfall[c].notna().any()]
    if candidate_columns:
        reference = rainfall[candidate_columns].median(axis=1, skipna=True)
    else:
        reference = pd.Series(np.nan, index=rainfall.index)
    network_wet = reference >= settings.network_wet_threshold_mm
    network_wet_hours = int(network_wet.sum())
    expected = len(rainfall)
    rows = []

    for station in rainfall.columns:
        series = rainfall[station]
        valid = series.dropna()
        included = True
        status = "active"
        reason = ""
        comparable = network_wet & series.notna()
        zero_when_wet_fraction = (
            float((series.loc[comparable] == 0).mean()) if comparable.any() else np.nan
        )

        if valid.empty:
            included = False
            status = "all_missing"
            reason = "No usable observations in the event period"
        elif (
            valid.nunique() == 1
            and len(valid) >= math.ceil(settings.constant_sensor_minimum_coverage * expected)
            and float(valid.iloc[0]) != 0.0
        ):
            included = False
            status = "constant_nonzero_value"
            reason = "Likely stuck sensor"
        elif (
            (valid > 0).sum() == 0
            and network_wet_hours >= settings.minimum_network_wet_hours
            and not pd.isna(zero_when_wet_fraction)
            and zero_when_wet_fraction >= settings.inactive_zero_fraction
        ):
            included = False
            status = "all_zero_while_network_wet"
            reason = "Likely inactive sensor"

        rows.append(
            {
                "event": event.name,
                "station": station,
                "included_in_analysis": included,
                "status": status,
                "reason": reason,
                "valid_observations": int(valid.count()),
                "expected_observations": expected,
                "positive_observations": int((valid > 0).sum()),
                "network_wet_hours": network_wet_hours,
                "zero_when_network_wet_fraction": zero_when_wet_fraction,
                "minimum_mm": float(valid.min()) if not valid.empty else np.nan,
                "maximum_mm": float(valid.max()) if not valid.empty else np.nan,
            }
        )
    return pd.DataFrame(rows)


def reject_obvious_network_outliers(
    event: EventConfig,
    rainfall: pd.DataFrame,
    active_stations: list[str],
    settings: QCSettings,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Remove only extremely large, isolated network outliers.

    This is intentionally stricter than a normal review flag. Values need to be
    >= obvious_outlier_min_mm, exceed the network median by both a ratio and an
    absolute difference, and be extreme by either MAD or a 6-IQR-style fence.
    """
    cleaned = rainfall.copy()
    rejected: list[dict[str, Any]] = []

    for timestamp, row in cleaned[active_stations].iterrows():
        available = row.dropna()
        if len(available) < settings.minimum_reporting_sensors_for_outlier_test:
            continue
        for station, value in available.items():
            if value < settings.obvious_outlier_min_mm:
                continue
            reference = available.drop(index=station)
            if len(reference) < settings.minimum_reporting_sensors_for_outlier_test - 1:
                continue
            metrics = robust_distribution_metrics(reference)
            median = metrics["median"]
            if pd.isna(median):
                continue
            difference = float(value - median)
            ratio = float((value + 0.5) / (median + 0.5))
            z = modified_z(float(value), median, metrics["mad"])
            iqr_upper = metrics["q3"] + settings.obvious_outlier_iqr_multiplier * metrics["iqr"]

            core = (
                ratio >= settings.obvious_outlier_min_ratio_to_network_median
                and difference >= settings.obvious_outlier_min_difference_mm
            )
            robust_extreme = (
                (not pd.isna(z) and z >= settings.obvious_outlier_modified_z)
                or (not pd.isna(iqr_upper) and value > iqr_upper)
            )
            if core and robust_extreme:
                rejected.append(
                    {
                        "event": event.name,
                        "timestamp": timestamp,
                        "station": station,
                        "original_value": float(value),
                        "reason": "obvious_network_outlier",
                        "network_median_mm": median,
                        "modified_z": z,
                        "network_ratio": ratio,
                        "network_difference_mm": difference,
                        "iqr_extreme_upper_mm": iqr_upper,
                    }
                )
                cleaned.at[timestamp, station] = np.nan

    return cleaned, pd.DataFrame(rejected, columns=REJECTED_COLUMNS)


def calculate_hourly_rainfall(
    event: EventConfig,
    rainfall: pd.DataFrame,
    active_stations: list[str],
) -> pd.DataFrame:
    active = rainfall[active_stations]
    hourly = pd.DataFrame(
        {
            "Event": event.name,
            "Date & Time": active.index,
            "Mean Rainfall (mm/h)": active.mean(axis=1, skipna=True).values,
            "Median Rainfall (mm/h)": active.median(axis=1, skipna=True).values,
            "Maximum Station Rainfall (mm/h)": active.max(axis=1, skipna=True).values,
            "Reporting Rain Gauges": active.count(axis=1).values,
            "Active Rain Gauges": len(active_stations),
        }
    )
    hourly["Date"] = hourly["Date & Time"].dt.floor("D")
    hourly["Hour"] = hourly["Date & Time"].dt.hour
    hourly["Daily Peak Hour"] = False
    for day, group in hourly.groupby("Date", sort=True):
        valid = group["Mean Rainfall (mm/h)"].dropna()
        if valid.empty:
            continue
        peak = valid.max()
        mask = group["Mean Rainfall (mm/h)"].notna() & np.isclose(
            group["Mean Rainfall (mm/h)"], peak, rtol=0, atol=1e-10
        )
        hourly.loc[group.index[mask], "Daily Peak Hour"] = True
    return hourly


def calculate_daily_rainfall(
    event: EventConfig,
    rainfall: pd.DataFrame,
    active_stations: list[str],
    settings: QCSettings,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Calculate daily station totals from available valid observations.

    A station-day is not discarded merely because some hours are missing. The
    available valid hourly observations are summed, and that partial-day total
    contributes to the network daily statistic as long as at least
    ``minimum_valid_hours_for_daily_accumulation`` observations are available.
    Missing hours remain missing and are never filled with zero.
    """
    active = rainfall[active_stations]
    station_day_rows: list[dict[str, Any]] = []

    for day in pd.date_range(event.start_date, event.end_date, freq="D"):
        day_data = active.loc[day : day + pd.Timedelta(hours=23)]
        network_hourly = day_data.median(axis=1, skipna=True)
        network_wet = network_hourly >= settings.network_wet_threshold_mm

        for station in active_stations:
            series = day_data[station]
            valid_hours = int(series.count())
            missing_hours = int(series.isna().sum())
            missing_wet_hours = int((series.isna() & network_wet).sum())
            qualifies = valid_hours >= settings.minimum_valid_hours_for_daily_accumulation

            station_day_rows.append(
                {
                    "Event": event.name,
                    "Date": day,
                    "Station": station,
                    "Daily Accumulation (mm)": (
                        float(series.sum(skipna=True)) if qualifies else np.nan
                    ),
                    "Valid Hours": valid_hours,
                    "Missing Hours": missing_hours,
                    "Expected Hours": 24,
                    "Coverage (%)": 100.0 * valid_hours / 24.0,
                    "Missing Network-Wet Hours": missing_wet_hours,
                    "Qualifies for Daily Statistics": qualifies,
                    "Daily Total Basis": (
                        "available valid observations" if qualifies else "no valid observations"
                    ),
                }
            )

    station_day = pd.DataFrame(station_day_rows)
    daily_rows = []
    for day, group in station_day.groupby("Date", sort=True):
        qualifying = group.loc[
            group["Qualifies for Daily Statistics"], "Daily Accumulation (mm)"
        ].dropna()
        daily_rows.append(
            {
                "Event": event.name,
                "Date": day,
                "Mean Daily Accumulated Rainfall (mm)": (
                    float(qualifying.mean()) if not qualifying.empty else np.nan
                ),
                "Median Daily Accumulated Rainfall (mm)": (
                    float(qualifying.median()) if not qualifying.empty else np.nan
                ),
                "Qualifying Rain Gauges": int(len(qualifying)),
                "Active Rain Gauges": int(len(active_stations)),
                "Mean Station-Day Coverage (%)": (
                    float(group.loc[group["Qualifies for Daily Statistics"], "Coverage (%)"].mean())
                    if not qualifying.empty
                    else np.nan
                ),
            }
        )
    return station_day, pd.DataFrame(daily_rows)


@dataclass
class RainfallResult:
    event: EventConfig
    cleaned: pd.DataFrame
    sensor_status: pd.DataFrame
    rejected: pd.DataFrame
    hourly: pd.DataFrame
    station_day: pd.DataFrame
    daily: pd.DataFrame
    qc_metadata: dict[str, Any]


def process_rainfall_event(event: EventConfig, settings: QCSettings) -> RainfallResult:
    raw_cleaned, rejected_hard, metadata = load_rainfall_event(event, settings)
    status = identify_sensor_status(event, raw_cleaned, settings)
    active_stations = status.loc[status["included_in_analysis"], "station"].tolist()
    if not active_stations:
        raise ValueError(f"No active rainfall sensors remain after QC for {event.name}.")

    # Excluded sensors remain in the cleaned workbook for auditability, but are
    # blanked so that downstream accidental means cannot include them.
    cleaned = raw_cleaned.copy()
    inactive = [c for c in cleaned.columns if c not in active_stations]
    if inactive:
        cleaned.loc[:, inactive] = np.nan

    cleaned, rejected_network = reject_obvious_network_outliers(
        event, cleaned, active_stations, settings
    )
    rejected_parts = [frame for frame in (rejected_hard, rejected_network) if not frame.empty]
    rejected = (
        pd.concat(rejected_parts, ignore_index=True)
        if rejected_parts
        else pd.DataFrame(columns=REJECTED_COLUMNS)
    )
    hourly = calculate_hourly_rainfall(event, cleaned, active_stations)
    station_day, daily = calculate_daily_rainfall(
        event, cleaned, active_stations, settings
    )

    metadata.update(
        {
            "active_sensors": len(active_stations),
            "excluded_sensors": int((~status["included_in_analysis"]).sum()),
            "hard_or_malformed_rejections": int(len(rejected_hard)),
            "obvious_network_outlier_rejections": int(len(rejected_network)),
            "total_rejected_values": int(len(rejected)),
            "remaining_missing_cells": int(cleaned[active_stations].isna().sum().sum()),
        }
    )
    return RainfallResult(
        event=event,
        cleaned=cleaned,
        sensor_status=status,
        rejected=rejected,
        hourly=hourly,
        station_day=station_day,
        daily=daily,
        qc_metadata=metadata,
    )


# =============================================================================
# FLOOD REPORTS
# =============================================================================


def load_august_flood_reports(path: Path, sheet_name: str) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Flood incidents workbook not found:\n{path}")
    frame = pd.read_excel(path, sheet_name=sheet_name, dtype=object, engine="openpyxl")
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if frame.empty:
        raise ValueError(f"Worksheet '{sheet_name}' contains no data.")

    start_date_col = find_column(frame.columns, "FLOOD START DATE")
    start_time_col = find_column(frame.columns, "FLOOD START TIME")
    end_date_col = find_column(frame.columns, "FLOOD END DATE")
    end_time_col = find_column(frame.columns, "FLOOD END TIME")
    duration_col = find_column(frame.columns, "DURATION")

    if start_date_col is None or start_time_col is None:
        raise ValueError(
            f"Worksheet '{sheet_name}' must contain FLOOD START DATE and FLOOD START TIME."
        )

    frame = frame.copy()
    frame["_Flood Start Datetime"] = [
        combine_excel_date_time(d, t)
        for d, t in zip(frame[start_date_col], frame[start_time_col])
    ]

    if end_date_col is not None and end_time_col is not None:
        frame["_Flood End Datetime"] = [
            combine_excel_date_time(d, t)
            for d, t in zip(frame[end_date_col], frame[end_time_col])
        ]
    else:
        frame["_Flood End Datetime"] = pd.NaT

    explicit_duration = (
        pd.to_datetime(frame["_Flood End Datetime"], errors="coerce")
        - pd.to_datetime(frame["_Flood Start Datetime"], errors="coerce")
    ).dt.total_seconds() / 60.0
    explicit_duration = explicit_duration.where(explicit_duration >= 0)

    if duration_col is not None:
        parsed_duration = frame[duration_col].map(parse_duration_minutes)
    else:
        parsed_duration = pd.Series(np.nan, index=frame.index, dtype=float)

    # Prefer start/end timestamps because they directly support the subsidence-time
    # analysis. Fall back to the DURATION field when end time is absent/unusable.
    frame["_Flood Duration Minutes"] = explicit_duration.fillna(parsed_duration)

    # If the source has a valid duration but no parseable end datetime, reconstruct
    # an estimated end datetime from start + duration.
    missing_end = (
        frame["_Flood End Datetime"].isna()
        & frame["_Flood Start Datetime"].notna()
        & frame["_Flood Duration Minutes"].notna()
    )
    frame.loc[missing_end, "_Flood End Datetime"] = (
        pd.to_datetime(frame.loc[missing_end, "_Flood Start Datetime"])
        + pd.to_timedelta(frame.loc[missing_end, "_Flood Duration Minutes"], unit="m")
    )
    return frame


def flood_tables_for_event(
    event: EventConfig,
    flood_all: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, int]:
    invalid_start_rows = int(flood_all["_Flood Start Datetime"].isna().sum())
    valid = flood_all.dropna(subset=["_Flood Start Datetime"]).copy()
    mask = (
        (valid["_Flood Start Datetime"] >= event.start_timestamp)
        & (valid["_Flood Start Datetime"] <= event.end_timestamp + pd.Timedelta(minutes=59, seconds=59))
    )
    reports = valid.loc[mask].copy()
    reports["_Flood Start Hour"] = reports["_Flood Start Datetime"].dt.floor("h")

    duration_values = pd.to_numeric(reports["_Flood Duration Minutes"], errors="coerce")
    duration_values = duration_values.where(duration_values >= 0)
    reports["_Flood Duration Minutes"] = duration_values

    # Event-level upper-duration outliers. These are not excluded; they are
    # retained and shown explicitly so the hourly average remains interpretable.
    valid_duration = duration_values.dropna()
    if len(valid_duration) >= 4:
        q1 = float(valid_duration.quantile(0.25))
        q3 = float(valid_duration.quantile(0.75))
        iqr = q3 - q1
        duration_outlier_threshold = q3 + 1.5 * iqr
    else:
        duration_outlier_threshold = np.nan

    reports["_Duration Outlier Threshold (min)"] = duration_outlier_threshold
    reports["_Duration Outlier"] = (
        reports["_Flood Duration Minutes"].notna()
        & np.isfinite(duration_outlier_threshold)
        & (reports["_Flood Duration Minutes"] > duration_outlier_threshold)
    )

    hourly_rows: list[dict[str, Any]] = []
    for timestamp in event.full_hourly_index:
        group = reports.loc[reports["_Flood Start Hour"] == timestamp]
        durations = group["_Flood Duration Minutes"].dropna()
        starts = group.loc[group["_Flood Start Datetime"].notna(), "_Flood Start Datetime"]
        ends = group.loc[group["_Flood End Datetime"].notna(), "_Flood End Datetime"]

        # For stakeholder visualization, compute a "typical" hourly average from
        # non-outlier reports. The all-report mean is still preserved separately
        # in the workbook so no information is lost. This prevents a single
        # multi-hour or multi-day report from flattening every ordinary duration
        # bar in the plot.
        typical_group = group.loc[~group["_Duration Outlier"]].copy()
        typical_durations = typical_group["_Flood Duration Minutes"].dropna()
        typical_starts = typical_group.loc[
            typical_group["_Flood Start Datetime"].notna(), "_Flood Start Datetime"
        ]
        typical_ends = typical_group.loc[
            typical_group["_Flood End Datetime"].notna(), "_Flood End Datetime"
        ]

        hourly_rows.append(
            {
                "Date & Time": timestamp,
                "Flood Reports": int(len(group)),
                "Reports with Valid Duration": int(len(durations)),
                "Average Flood Start Datetime": mean_datetime(starts),
                "Average Flood Duration (min)": (float(durations.mean()) if not durations.empty else np.nan),
                "Median Flood Duration (min)": (float(durations.median()) if not durations.empty else np.nan),
                "Longest Flood Duration (min)": (float(durations.max()) if not durations.empty else np.nan),
                "Average Subsidence Datetime": mean_datetime(ends),
                "Typical Reports (Non-Outliers)": int(len(typical_durations)),
                "Average Typical Flood Duration (min)": (
                    float(typical_durations.mean()) if not typical_durations.empty else np.nan
                ),
                "Average Typical Flood Start Datetime": mean_datetime(typical_starts),
                "Average Typical Subsidence Datetime": mean_datetime(typical_ends),
                "Duration Outlier Reports": int(group["_Duration Outlier"].sum()) if not group.empty else 0,
            }
        )

    hourly = pd.DataFrame(hourly_rows)
    hourly["Date"] = hourly["Date & Time"].dt.floor("D")
    hourly["Hour"] = hourly["Date & Time"].dt.hour

    daily_rows = []
    for day, group in hourly.groupby("Date", sort=True):
        peak = int(group["Flood Reports"].max())
        peak_rows = group.loc[group["Flood Reports"] == peak]
        peak_time = peak_rows.iloc[0]["Date & Time"] if peak > 0 else pd.NaT

        source_day = reports.loc[reports["_Flood Start Datetime"].dt.floor("D") == day]
        day_durations = source_day["_Flood Duration Minutes"].dropna()
        daily_rows.append(
            {
                "Date": day,
                "Total Flood Reports": int(group["Flood Reports"].sum()),
                "Peak Hourly Flood Reports": peak,
                "Peak Flood Report Hour": peak_time,
                "Average Flood Duration (min)": (float(day_durations.mean()) if not day_durations.empty else np.nan),
                "Median Flood Duration (min)": (float(day_durations.median()) if not day_durations.empty else np.nan),
                "Longest Flood Duration (min)": (float(day_durations.max()) if not day_durations.empty else np.nan),
                "Duration Outlier Reports": int(source_day["_Duration Outlier"].sum()) if not source_day.empty else 0,
            }
        )
    daily = pd.DataFrame(daily_rows)
    return hourly, daily, reports, invalid_start_rows


# =============================================================================
# SUMMARY
# =============================================================================


def build_event_summary(
    rainfall: RainfallResult,
    flood_hourly: pd.DataFrame,
    flood_daily: pd.DataFrame,
    invalid_flood_start_rows: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    event = rainfall.event
    daily = rainfall.daily.merge(flood_daily, on="Date", how="left", validate="one_to_one")

    # Add daily rainfall peak hour/value.
    peak_rain_rows = []
    for day, group in rainfall.hourly.groupby("Date", sort=True):
        valid = group.dropna(subset=["Mean Rainfall (mm/h)"])
        if valid.empty:
            peak_time = pd.NaT
            peak_value = np.nan
        else:
            peak_idx = valid["Mean Rainfall (mm/h)"].idxmax()
            peak_time = valid.loc[peak_idx, "Date & Time"]
            peak_value = float(valid.loc[peak_idx, "Mean Rainfall (mm/h)"])
        peak_rain_rows.append(
            {
                "Date": day,
                "Peak Hourly Mean Rainfall (mm/h)": peak_value,
                "Peak Rainfall Hour": peak_time,
            }
        )
    daily = daily.merge(pd.DataFrame(peak_rain_rows), on="Date", how="left", validate="one_to_one")

    daily = daily[
        [
            "Date",
            "Mean Daily Accumulated Rainfall (mm)",
            "Median Daily Accumulated Rainfall (mm)",
            "Qualifying Rain Gauges",
            "Active Rain Gauges",
            "Mean Station-Day Coverage (%)",
            "Peak Hourly Mean Rainfall (mm/h)",
            "Peak Rainfall Hour",
            "Total Flood Reports",
            "Peak Hourly Flood Reports",
            "Peak Flood Report Hour",
        ]
    ]

    valid_rain_hours = rainfall.hourly.dropna(subset=["Mean Rainfall (mm/h)"])
    if valid_rain_hours.empty:
        event_peak_rain = np.nan
        event_peak_rain_time = pd.NaT
    else:
        idx = valid_rain_hours["Mean Rainfall (mm/h)"].idxmax()
        event_peak_rain = float(valid_rain_hours.loc[idx, "Mean Rainfall (mm/h)"])
        event_peak_rain_time = valid_rain_hours.loc[idx, "Date & Time"]

    max_flood = int(flood_hourly["Flood Reports"].max()) if not flood_hourly.empty else 0
    if max_flood > 0:
        idx = flood_hourly["Flood Reports"].idxmax()
        event_peak_flood_time = flood_hourly.loc[idx, "Date & Time"]
    else:
        event_peak_flood_time = pd.NaT

    event_summary = pd.DataFrame(
        [
            {
                "Event": event.name,
                "Period": event.label,
                "Event Accumulated Rainfall (mm)": daily[
                    "Mean Daily Accumulated Rainfall (mm)"
                ].sum(min_count=1),
                "Peak Hourly Mean Rainfall (mm/h)": event_peak_rain,
                "Peak Rainfall Hour": event_peak_rain_time,
                "Total Flood Reports": int(flood_hourly["Flood Reports"].sum()),
                "Peak Hourly Flood Reports": max_flood,
                "Peak Flood Report Hour": event_peak_flood_time,
                "Input Rain Gauges": rainfall.qc_metadata["input_station_columns"],
                "Active Rain Gauges": rainfall.qc_metadata["active_sensors"],
                "Excluded Rain Gauges": rainfall.qc_metadata["excluded_sensors"],
                "Rejected Rainfall Values": rainfall.qc_metadata["total_rejected_values"],
                "Obvious Network Outliers Rejected": rainfall.qc_metadata[
                    "obvious_network_outlier_rejections"
                ],
                "Remaining Missing Rainfall Cells": rainfall.qc_metadata[
                    "remaining_missing_cells"
                ],
                "Flood Rows with Unparseable Start Datetime in Source Sheet": invalid_flood_start_rows,
            }
        ]
    )
    return daily, event_summary


def format_timestamp(value: Any) -> str:
    if pd.isna(value):
        return "N/A"
    return pd.Timestamp(value).strftime("%d %b %Y %H:%M")


def write_plain_text_summary(
    path: Path,
    daily_summary: pd.DataFrame,
    event_summary: pd.DataFrame,
) -> None:
    row = event_summary.iloc[0]
    lines = [
        f"{row['Event']} ({row['Period']})",
        "=" * 72,
        "",
        "Rainfall metric notes:",
        "- Hourly rainfall = mean across valid active rain gauges reporting that hour.",
        "- Daily accumulation = mean of station daily sums from available valid observations.",
        "- A station-day may contribute even when some hours are missing; missing hours are skipped, not filled with zero.",
        "- Event accumulation = sum of available daily mean accumulations.",
        "- Missing values and rejected outliers are not treated as zero rainfall.",
        "",
        "DAILY SUMMARY",
    ]
    for _, day in daily_summary.iterrows():
        lines.extend(
            [
                f"{pd.Timestamp(day['Date']):%d %b %Y}",
                f"  Daily accumulated rainfall: {day['Mean Daily Accumulated Rainfall (mm)']:.2f} mm"
                if not pd.isna(day["Mean Daily Accumulated Rainfall (mm)"])
                else "  Daily accumulated rainfall: N/A",
                (
                    f"  Peak hourly mean rainfall: {day['Peak Hourly Mean Rainfall (mm/h)']:.2f} mm/h "
                    f"at {format_timestamp(day['Peak Rainfall Hour'])}"
                    if not pd.isna(day["Peak Hourly Mean Rainfall (mm/h)"])
                    else "  Peak hourly mean rainfall: N/A"
                ),
                f"  Total flood reports: {int(day['Total Flood Reports'])}",
                (
                    f"  Peak hourly flood reports: {int(day['Peak Hourly Flood Reports'])} "
                    f"at {format_timestamp(day['Peak Flood Report Hour'])}"
                    if int(day["Peak Hourly Flood Reports"]) > 0
                    else "  Peak hourly flood reports: 0"
                ),
                "",
            ]
        )

    lines.extend(
        [
            "WHOLE EVENT",
            f"Event accumulated rainfall: {row['Event Accumulated Rainfall (mm)']:.2f} mm"
            if not pd.isna(row["Event Accumulated Rainfall (mm)"])
            else "Event accumulated rainfall: N/A",
            (
                f"Peak hourly mean rainfall: {row['Peak Hourly Mean Rainfall (mm/h)']:.2f} mm/h "
                f"at {format_timestamp(row['Peak Rainfall Hour'])}"
                if not pd.isna(row["Peak Hourly Mean Rainfall (mm/h)"])
                else "Peak hourly mean rainfall: N/A"
            ),
            f"Total flood reports: {int(row['Total Flood Reports'])}",
            (
                f"Peak hourly flood reports: {int(row['Peak Hourly Flood Reports'])} "
                f"at {format_timestamp(row['Peak Flood Report Hour'])}"
                if int(row["Peak Hourly Flood Reports"]) > 0
                else "Peak hourly flood reports: 0"
            ),
            "",
            "QC",
            f"Active rain gauges: {int(row['Active Rain Gauges'])} of {int(row['Input Rain Gauges'])}",
            f"Excluded rain gauges: {int(row['Excluded Rain Gauges'])}",
            f"Rejected rainfall values: {int(row['Rejected Rainfall Values'])}",
            f"  of which obvious network outliers: {int(row['Obvious Network Outliers Rejected'])}",
            f"Remaining missing rainfall cells: {int(row['Remaining Missing Rainfall Cells'])}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# =============================================================================
# PLOTTING
# =============================================================================


def apply_plot_style() -> None:
    plt.rcParams.update(
        {
            "figure.dpi": 130,
            "savefig.dpi": 190,
            "font.size": 10,
            "axes.titlesize": 13,
            "axes.labelsize": 10,
            "axes.spines.top": False,
            "axes.spines.right": False,
        }
    )


def add_intensity_lines(ax: plt.Axes) -> None:
    for threshold, label, color in HOURLY_INTENSITY_LINES:
        ax.axhline(
            threshold,
            color=color,
            linestyle="--",
            linewidth=1.15,
            alpha=0.95,
            label=f"{label} ({threshold:g} mm/h)",
        )


def time_axis_ticks(timestamps: pd.Series, step_hours: int = 6) -> tuple[np.ndarray, list[str]]:
    x = np.arange(len(timestamps))
    tick_positions = x[::step_hours]
    labels = [pd.Timestamp(timestamps.iloc[i]).strftime("%d %b\n%H:%M") for i in tick_positions]
    return tick_positions, labels


def finish_figure(fig: plt.Figure, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_hourly_rainfall_3day(
    result: RainfallResult,
    output_path: Path,
    rainfall_ymax: float,
) -> None:
    hourly = result.hourly.sort_values("Date & Time").reset_index(drop=True)
    x = np.arange(len(hourly))
    values = hourly["Mean Rainfall (mm/h)"].to_numpy(float)

    fig, ax = plt.subplots(figsize=(18, 6.2))
    bars = ax.bar(x, values, width=0.82, color=result.event.color, label="Hourly network-mean rainfall")

    peak_mask = hourly["Daily Peak Hour"].to_numpy(bool)
    peak_indices = np.where(peak_mask)[0]
    for idx in peak_indices:
        bars[idx].set_facecolor("#FFD700")
        bars[idx].set_edgecolor("0.25")
        bars[idx].set_linewidth(0.8)

    add_intensity_lines(ax)
    peak_patch = Patch(
        facecolor="#FFD700",
        edgecolor="0.25",
        linewidth=0.8,
        label="Daily peak rainfall hour",
    )
    handles, labels = ax.get_legend_handles_labels()
    handles.append(peak_patch)
    labels.append("Daily peak rainfall hour")

    ticks, tick_labels = time_axis_ticks(hourly["Date & Time"], step_hours=6)
    ax.set_xticks(ticks)
    ax.set_xticklabels(tick_labels)
    ax.set_xlim(-0.8, len(hourly) - 0.2)
    ax.set_ylim(0, rainfall_ymax)
    ax.set_xlabel("Date and time (PHT)")
    ax.set_ylabel("Network-mean rainfall (mm/h)")
    ax.set_title(f"{result.event.label}: Hourly Rainfall Across the 3-Day Event")

    # Day separators at midnight between event days.
    for boundary in (24, 48):
        ax.axvline(boundary - 0.5, color="0.55", linewidth=0.9, linestyle=":", alpha=0.8)

    ax.legend(handles, labels, ncol=3, fontsize=8, frameon=True, loc="upper right")
    finish_figure(fig, output_path)


def plot_hourly_rainfall_and_flood(
    result: RainfallResult,
    flood_hourly: pd.DataFrame,
    output_path: Path,
    rainfall_ymax: float,
    flood_ymax: float,
) -> None:
    hourly = result.hourly.sort_values("Date & Time").reset_index(drop=True)
    flood = flood_hourly.sort_values("Date & Time").reset_index(drop=True)
    if not hourly["Date & Time"].equals(flood["Date & Time"]):
        raise ValueError(f"Rainfall and flood timelines do not align for {result.event.name}.")

    x = np.arange(len(hourly))
    rainfall_values = hourly["Mean Rainfall (mm/h)"].to_numpy(float)
    flood_values = flood["Flood Reports"].to_numpy(int)

    fig, ax1 = plt.subplots(figsize=(18, 6.5))
    bars = ax1.bar(
        x,
        rainfall_values,
        width=0.82,
        color=result.event.color,
        alpha=0.82,
        label="Hourly network-mean rainfall",
        zorder=2,
    )
    peak_mask = hourly["Daily Peak Hour"].to_numpy(bool)
    for idx in np.where(peak_mask)[0]:
        bars[idx].set_facecolor("#FFD700")
        bars[idx].set_edgecolor("0.25")
        bars[idx].set_linewidth(0.8)

    add_intensity_lines(ax1)
    ax1.set_ylim(0, rainfall_ymax)
    ax1.set_ylabel("Network-mean rainfall (mm/h)")
    ax1.set_xlabel("Date and time (PHT)")

    ax2 = ax1.twinx()
    ax2.plot(
        x,
        flood_values,
        color="black",
        marker="o",
        markersize=3.8,
        linewidth=1.6,
        label="Flood reports starting in hour",
        zorder=5,
    )
    ax2.set_ylim(0, flood_ymax)
    ax2.set_ylabel("Flood reports starting in hour")
    ax2.yaxis.set_major_locator(MaxNLocator(integer=True, nbins=7))
    ax2.spines["right"].set_visible(True)

    ticks, tick_labels = time_axis_ticks(hourly["Date & Time"], step_hours=6)
    ax1.set_xticks(ticks)
    ax1.set_xticklabels(tick_labels)
    ax1.set_xlim(-0.8, len(hourly) - 0.2)
    ax1.set_title(f"{result.event.label}: Hourly Rainfall and Flood Reports")

    for boundary in (24, 48):
        ax1.axvline(boundary - 0.5, color="0.55", linewidth=0.9, linestyle=":", alpha=0.8)

    peak_patch = Patch(
        facecolor="#FFD700",
        edgecolor="0.25",
        linewidth=0.8,
        label="Daily peak rainfall hour",
    )
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + [peak_patch] + h2, l1 + ["Daily peak rainfall hour"] + l2, ncol=3, fontsize=8, loc="upper right")
    finish_figure(fig, output_path)


def plot_hourly_rainfall_and_flood_duration(
    result: RainfallResult,
    flood_hourly: pd.DataFrame,
    flood_reports: pd.DataFrame,
    output_path: Path,
    rainfall_ymax: float,
    duration_xmax: float,
) -> None:
    """Rainfall context plus stakeholder-friendly horizontal duration bars.

    The top panel preserves the 72-hour rainfall chronology. The lower panel is
    intentionally a separate horizontal bar chart because duration is not itself
    a continuous hourly time series. Each row is an hour in which flooding
    started. The black bar is the mean duration of non-outlier reports from that
    start hour; the right-side row label gives their average subsidence time.

    Long-duration IQR outliers are retained and shown in red. If an outlier lies
    beyond the common duration axis, a red right-pointing triangle is placed at
    the axis edge. Full outlier durations remain available in the Excel output.
    """
    hourly = result.hourly.sort_values("Date & Time").reset_index(drop=True)
    flood = flood_hourly.sort_values("Date & Time").reset_index(drop=True)
    if not hourly["Date & Time"].equals(flood["Date & Time"]):
        raise ValueError(f"Rainfall and flood timelines do not align for {result.event.name}.")

    # Include every hour with at least one valid flood duration, even if all of
    # that hour's reports are outliers (in which case there is no black bar).
    duration_rows = flood.loc[flood["Reports with Valid Duration"] > 0].copy()
    duration_rows = duration_rows.sort_values("Date & Time").reset_index(drop=True)

    n_rows = len(duration_rows)
    fig_height = max(8.0, min(15.5, 6.4 + 0.26 * n_rows))
    fig, (ax_rain, ax_duration) = plt.subplots(
        2,
        1,
        figsize=(18, fig_height),
        gridspec_kw={"height_ratios": [1.12, max(1.0, min(2.4, 0.075 * max(n_rows, 1) + 0.9))], "hspace": 0.18},
    )

    # ---------------------------------------------------------------------
    # TOP: hourly rainfall context (same style as the main rainfall figure)
    # ---------------------------------------------------------------------
    x = np.arange(len(hourly))
    rainfall_values = hourly["Mean Rainfall (mm/h)"].to_numpy(float)
    bars = ax_rain.bar(
        x,
        rainfall_values,
        width=0.82,
        color=result.event.color,
        alpha=0.82,
        label="Hourly network-mean rainfall",
        zorder=2,
    )
    peak_mask = hourly["Daily Peak Hour"].to_numpy(bool)
    for idx in np.where(peak_mask)[0]:
        bars[idx].set_facecolor("#FFD700")
        bars[idx].set_edgecolor("0.25")
        bars[idx].set_linewidth(0.8)

    add_intensity_lines(ax_rain)
    ax_rain.set_ylim(0, rainfall_ymax)
    ax_rain.set_ylabel("Network-mean rainfall (mm/h)")
    ax_rain.set_xlabel("Date and time (PHT)")
    ax_rain.set_title(f"{result.event.label}: Hourly Rainfall and Flood Duration")
    ticks, tick_labels = time_axis_ticks(hourly["Date & Time"], step_hours=6)
    ax_rain.set_xticks(ticks)
    ax_rain.set_xticklabels(tick_labels)
    ax_rain.set_xlim(-0.8, len(hourly) - 0.2)
    for boundary in (24, 48):
        ax_rain.axvline(boundary - 0.5, color="0.55", linewidth=0.9, linestyle=":", alpha=0.8)

    peak_patch = Patch(
        facecolor="#FFD700",
        edgecolor="0.25",
        linewidth=0.8,
        label="Daily peak rainfall hour",
    )
    h1, l1 = ax_rain.get_legend_handles_labels()
    ax_rain.legend(
        h1 + [peak_patch],
        l1 + ["Daily peak rainfall hour"],
        ncol=3,
        fontsize=8,
        loc="upper right",
    )

    # ---------------------------------------------------------------------
    # BOTTOM: horizontal duration bars by flood start hour
    # ---------------------------------------------------------------------
    if duration_rows.empty:
        ax_duration.text(
            0.5,
            0.5,
            "No flood reports with usable duration data",
            transform=ax_duration.transAxes,
            ha="center",
            va="center",
            fontsize=10,
        )
        ax_duration.set_yticks([])
        ax_duration.set_xlim(0, duration_xmax)
    else:
        y = np.arange(len(duration_rows))
        typical = pd.to_numeric(
            duration_rows["Average Typical Flood Duration (min)"], errors="coerce"
        ).to_numpy(float)

        ax_duration.barh(
            y,
            np.nan_to_num(typical, nan=0.0),
            height=0.52,
            color="black",
            alpha=0.78,
            label="Average duration (non-outlier reports)",
            zorder=3,
        )

        left_labels = []
        right_labels = []
        outlier_label_added = False
        overflow_label_added = False

        for row_index, row in duration_rows.iterrows():
            start_hour = pd.Timestamp(row["Date & Time"])
            n_valid = int(row["Reports with Valid Duration"])
            n_outliers = int(row["Duration Outlier Reports"])
            left_labels.append(
                f"{start_hour.strftime('%d %b %H:00')}  (n={n_valid}, out={n_outliers})"
            )

            typical_end = row.get("Average Typical Subsidence Datetime", pd.NaT)
            if pd.notna(typical_end):
                end_ts = pd.Timestamp(typical_end)
                # Include the date only when subsidence crossed into another day.
                if end_ts.date() == start_hour.date():
                    right_labels.append(end_ts.strftime("avg subsides %H:%M"))
                else:
                    right_labels.append(end_ts.strftime("avg subsides %d %b %H:%M"))
            else:
                right_labels.append("all duration reports are outliers")

            group_outliers = flood_reports.loc[
                flood_reports.get("_Duration Outlier", False)
                & flood_reports["_Flood Start Datetime"].notna()
                & flood_reports["_Flood Duration Minutes"].notna()
                & (pd.to_datetime(flood_reports["_Flood Start Datetime"]).dt.floor("h") == start_hour)
            ].copy()

            if group_outliers.empty:
                continue

            values = pd.to_numeric(
                group_outliers["_Flood Duration Minutes"], errors="coerce"
            ).dropna().to_numpy(float)
            if len(values) == 0:
                continue

            offsets = np.linspace(-0.19, 0.19, len(values)) if len(values) > 1 else np.array([0.0])
            for offset, value in zip(offsets, values):
                if value <= duration_xmax:
                    ax_duration.scatter(
                        value,
                        row_index + float(offset),
                        color="#C00000",
                        marker="D",
                        s=34,
                        edgecolors="white",
                        linewidths=0.5,
                        label=("Long-duration report (IQR outlier)" if not outlier_label_added else None),
                        zorder=6,
                    )
                    outlier_label_added = True
                else:
                    ax_duration.scatter(
                        duration_xmax * 0.985,
                        row_index + float(offset),
                        color="#C00000",
                        marker=">",
                        s=56,
                        label=("Outlier extends beyond displayed scale" if not overflow_label_added else None),
                        zorder=6,
                        clip_on=False,
                    )
                    overflow_label_added = True

        ax_duration.set_yticks(y)
        ax_duration.set_yticklabels(left_labels, fontsize=8)
        ax_duration.invert_yaxis()
        ax_duration.set_xlim(0, duration_xmax)
        ax_duration.set_xlabel("Flood duration (minutes)")
        ax_duration.set_ylabel("Flood start hour")
        ax_duration.grid(axis="x", linestyle=":", linewidth=0.6, alpha=0.5)
        ax_duration.set_title(
            "Average flood duration by start hour; long-duration reports highlighted in red",
            loc="left",
            fontsize=10,
            pad=6,
        )

        # A right-hand label column directly answers the operational question:
        # around what time did floods starting in this hour usually subside?
        ax_right = ax_duration.twinx()
        ax_right.set_ylim(ax_duration.get_ylim())
        ax_right.set_yticks(y)
        ax_right.set_yticklabels(right_labels, fontsize=8)
        ax_right.tick_params(axis="y", length=0, pad=8)
        ax_right.set_ylabel("Average subsidence time (non-outliers)")
        ax_right.spines["right"].set_visible(False)

        handles, labels = ax_duration.get_legend_handles_labels()
        ax_duration.legend(handles, labels, loc="upper right", fontsize=8, frameon=True)

    finish_figure(fig, output_path)


# =============================================================================
# OUTPUT WRITING
# =============================================================================


def write_cleaned_rainfall_workbook(result: RainfallResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cleaned = result.cleaned.copy().reset_index()
    cleaned.insert(1, "Network Mean Rainfall (mm/h)", result.hourly["Mean Rainfall (mm/h)"].values)
    cleaned.insert(2, "Reporting Rain Gauges", result.hourly["Reporting Rain Gauges"].values)

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm") as writer:
        cleaned.to_excel(writer, sheet_name="Cleaned Rainfall", index=False)
        result.hourly.to_excel(writer, sheet_name="Hourly Network Summary", index=False)
        result.station_day.to_excel(writer, sheet_name="Station Daily QC", index=False)
        result.daily.to_excel(writer, sheet_name="Daily Rainfall Summary", index=False)
    format_excel_workbook(path)


def write_flood_workbook(
    hourly: pd.DataFrame,
    daily: pd.DataFrame,
    reports: pd.DataFrame,
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    hourly_export = hourly[
        [
            "Date & Time",
            "Flood Reports",
            "Reports with Valid Duration",
            "Average Flood Start Datetime",
            "Average Flood Duration (min)",
            "Median Flood Duration (min)",
            "Longest Flood Duration (min)",
            "Average Subsidence Datetime",
            "Typical Reports (Non-Outliers)",
            "Average Typical Flood Duration (min)",
            "Average Typical Flood Start Datetime",
            "Average Typical Subsidence Datetime",
            "Duration Outlier Reports",
        ]
    ].copy()

    reports_export = reports.copy()
    reports_export = reports_export.drop(columns=["_Flood Start Hour"], errors="ignore")
    outlier_reports = reports_export.loc[reports_export["_Duration Outlier"]].copy()

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm") as writer:
        hourly_export[["Date & Time", "Flood Reports"]].to_excel(
            writer, sheet_name="Hourly Flood Reports", index=False
        )
        hourly_export.to_excel(writer, sheet_name="Hourly Flood Duration", index=False)

        # One 24-row sheet per day, now including duration/subsidence metrics.
        for day, group in hourly.groupby("Date", sort=True):
            day_table = group[
                [
                    "Date & Time",
                    "Flood Reports",
                    "Reports with Valid Duration",
                    "Average Flood Start Datetime",
                    "Average Flood Duration (min)",
                    "Median Flood Duration (min)",
                    "Longest Flood Duration (min)",
                    "Average Subsidence Datetime",
                    "Typical Reports (Non-Outliers)",
                    "Average Typical Flood Duration (min)",
                    "Average Typical Flood Start Datetime",
                    "Average Typical Subsidence Datetime",
                    "Duration Outlier Reports",
                ]
            ].copy()
            day_table.to_excel(
                writer,
                sheet_name=safe_excel_sheet_name(pd.Timestamp(day).strftime("%d %b")),
                index=False,
            )

        daily.to_excel(writer, sheet_name="Daily Flood Summary", index=False)
        outlier_reports.to_excel(writer, sheet_name="Duration Outliers", index=False)
        reports_export.to_excel(writer, sheet_name="Source Reports in Event", index=False)
    format_excel_workbook(path)


def write_summary_workbook(
    daily_summary: pd.DataFrame,
    event_summary: pd.DataFrame,
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm") as writer:
        daily_summary.to_excel(writer, sheet_name="Daily Summary", index=False)
        event_summary.to_excel(writer, sheet_name="Event Summary", index=False)
    format_excel_workbook(path)


def write_qc_summary(result: RainfallResult, path: Path, settings: QCSettings) -> None:
    lines = [
        f"QC summary: {result.event.name}",
        "=" * 72,
        "",
        "Automatic rejection rules:",
        f"- Negative rainfall: rejected",
        f"- Hard hourly ceiling: > {settings.hard_max_hourly_mm:g} mm/h rejected",
        (
            "- Additional network outlier: value must be >= "
            f"{settings.obvious_outlier_min_mm:g} mm/h, >= "
            f"{settings.obvious_outlier_min_ratio_to_network_median:g}x network median, "
            f">= {settings.obvious_outlier_min_difference_mm:g} mm above network median, "
            "and extremely unusual by robust MAD/IQR statistics"
        ),
        "- Missing values remain missing and are excluded; they are never replaced with zero.",
        "",
    ]
    for key, value in result.qc_metadata.items():
        lines.append(f"{key}: {value}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_event_outputs(
    result: RainfallResult,
    flood_hourly: pd.DataFrame,
    flood_daily: pd.DataFrame,
    flood_reports: pd.DataFrame,
    daily_summary: pd.DataFrame,
    event_summary: pd.DataFrame,
    invalid_flood_start_rows: int,
    root: Path,
    settings: QCSettings,
    rainfall_ymax: float,
    flood_ymax: float,
    duration_xmax: float,
) -> None:
    event_root = root / result.event.name
    plots = event_root / "plots"
    tables = event_root / "tables"
    qc = event_root / "qc"
    for folder in (plots, tables, qc):
        folder.mkdir(parents=True, exist_ok=True)

    write_cleaned_rainfall_workbook(result, tables / "cleaned_rainfall_hourly.xlsx")
    write_flood_workbook(flood_hourly, flood_daily, flood_reports, tables / "flood_reports_hourly.xlsx")
    write_summary_workbook(daily_summary, event_summary, tables / "event_summary.xlsx")

    save_csv(result.rejected, qc / "rejected_rainfall_values.csv")
    save_csv(result.sensor_status, qc / "sensor_status.csv")
    write_qc_summary(result, qc / "qc_summary.txt", settings)
    write_plain_text_summary(event_root / "event_summary.txt", daily_summary, event_summary)

    plot_hourly_rainfall_3day(
        result,
        plots / "hourly_rainfall_3day.png",
        rainfall_ymax,
    )
    plot_hourly_rainfall_and_flood(
        result,
        flood_hourly,
        plots / "hourly_rainfall_and_flood_reports_3day.png",
        rainfall_ymax,
        flood_ymax,
    )
    plot_hourly_rainfall_and_flood_duration(
        result,
        flood_hourly,
        flood_reports,
        plots / "hourly_rainfall_and_flood_duration_3day.png",
        rainfall_ymax,
        duration_xmax,
    )


# =============================================================================
# ANALYSIS ORCHESTRATION
# =============================================================================


def run_analysis(
    events: tuple[EventConfig, ...],
    flood_file: Path,
    flood_sheet: str,
    output_root: Path,
    settings: QCSettings,
) -> None:
    output_root.mkdir(parents=True, exist_ok=True)

    # Rainfall processing is independent for the two events and is intentionally
    # run in parallel. Plotting waits until both complete so common axis limits
    # can be derived from both events together.
    rainfall_results: dict[str, RainfallResult] = {}
    with ThreadPoolExecutor(max_workers=len(events), thread_name_prefix="AugHab") as executor:
        futures = {executor.submit(process_rainfall_event, event, settings): event for event in events}
        for future in as_completed(futures):
            event = futures[future]
            rainfall_results[event.name] = future.result()
            print(f"Rainfall QC complete: {event.name}")

    flood_all = load_august_flood_reports(flood_file, flood_sheet)
    flood_results: dict[str, tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, int]] = {}
    for event in events:
        flood_results[event.name] = flood_tables_for_event(event, flood_all)

    # Shared rainfall axis across BOTH rainfall plots and BOTH combination plots.
    all_hourly_rain = pd.concat(
        [rainfall_results[e.name].hourly["Mean Rainfall (mm/h)"] for e in events],
        ignore_index=True,
    )
    rainfall_ymax = common_axis_upper(
        all_hourly_rain,
        minimum_peak=max(v for v, _, _ in HOURLY_INTENSITY_LINES),
        padding=1.12,
    )

    # Shared flood-report axis across BOTH combination plots.
    all_flood_counts = pd.concat(
        [flood_results[e.name][0]["Flood Reports"] for e in events], ignore_index=True
    )
    flood_ymax = common_axis_upper(all_flood_counts, minimum_peak=1.0, padding=1.18)
    flood_ymax = max(1.0, math.ceil(flood_ymax))

    # Shared duration axis for the stakeholder horizontal-bar plots. Deliberately
    # derive this from NON-OUTLIER durations so very long reports do not compress
    # the ordinary bars. Outliers beyond the scale are shown with red > markers.
    typical_duration_values: list[float] = []
    for event in events:
        reports = flood_results[event.name][2]
        typical_report_durations = pd.to_numeric(
            reports.loc[~reports["_Duration Outlier"], "_Flood Duration Minutes"],
            errors="coerce",
        ).dropna()
        typical_duration_values.extend(typical_report_durations.tolist())
        hourly_typical = pd.to_numeric(
            flood_results[event.name][0]["Average Typical Flood Duration (min)"],
            errors="coerce",
        ).dropna()
        typical_duration_values.extend(hourly_typical.tolist())

    duration_xmax = common_axis_upper(
        typical_duration_values,
        minimum_peak=60.0,
        padding=1.15,
    )
    # Round to a stakeholder-friendly 30-minute boundary.
    duration_xmax = max(60.0, math.ceil(duration_xmax / 30.0) * 30.0)

    apply_plot_style()
    for event in events:
        result = rainfall_results[event.name]
        flood_hourly, flood_daily, flood_reports, invalid_flood_start_rows = flood_results[event.name]
        daily_summary, event_summary = build_event_summary(
            result, flood_hourly, flood_daily, invalid_flood_start_rows
        )
        write_event_outputs(
            result,
            flood_hourly,
            flood_daily,
            flood_reports,
            daily_summary,
            event_summary,
            invalid_flood_start_rows,
            output_root,
            settings,
            rainfall_ymax,
            flood_ymax,
            duration_xmax,
        )
        print(f"Outputs written: {output_root / event.name}")

    print("\nAnalysis complete.")
    print(f"Common rainfall y-axis: 0 to {rainfall_ymax:g} mm/h")
    print(f"Common flood-report y-axis: 0 to {flood_ymax:g} reports/hour")
    print(f"Common typical-duration x-axis: 0 to {duration_xmax:g} minutes")
    print(f"Outputs: {output_root}")


# =============================================================================
# CLI
# =============================================================================


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Analyze 08-10 Aug and 17-19 Aug 2026 rainfall/flooding events."
    )
    parser.add_argument("--rf-0810", type=Path, default=RF_0810_FILE)
    parser.add_argument("--rf-1719", type=Path, default=RF_1719_FILE)
    parser.add_argument("--flood-reports", type=Path, default=FLOOD_REPORTS_FILE)
    parser.add_argument("--flood-sheet", default=FLOOD_SHEET)
    parser.add_argument("--output", type=Path, default=OUTPUT_DIRECTORY)
    return parser


def main() -> int:
    args = build_argument_parser().parse_args()
    events = (
        EventConfig(
            name="Aug08-10",
            label="08-10 Aug 2026",
            rainfall_file=args.rf_0810,
            start_date=pd.Timestamp("2026-08-08"),
            end_date=pd.Timestamp("2026-08-10"),
            color="#ae8df5",
        ),
        EventConfig(
            name="Aug17-19",
            label="17-19 Aug 2026",
            rainfall_file=args.rf_1719,
            start_date=pd.Timestamp("2026-08-17"),
            end_date=pd.Timestamp("2026-08-19"),
            color="#f3b86b",
        ),
    )

    try:
        run_analysis(
            events=events,
            flood_file=args.flood_reports,
            flood_sheet=args.flood_sheet,
            output_root=args.output,
            settings=DEFAULT_QC,
        )
    except (FileNotFoundError, ValueError, RuntimeError, KeyError) as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
